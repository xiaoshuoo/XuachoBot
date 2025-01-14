import os
import logging
import json
import sqlite3
from datetime import datetime
from aiogram import Bot, Dispatcher, types, F, Router
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    InlineKeyboardButton, 
    InlineKeyboardMarkup, 
    FSInputFile, 
    Message, 
    ContentType,
    LinkPreviewOptions  # Добавляем импорт
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.enums import ParseMode
from aiogram.utils.chat_action import ChatActionMiddleware, ChatActionSender
from aiogram.exceptions import (
    TelegramNetworkError, 
    TelegramBadRequest,
    TelegramAPIError
)
from aiogram.utils.formatting import (
    Text,
    Bold,
    Code,
    Pre,
    Italic,
    TextLink,
    Underline
)
import google.generativeai as genai
from aiogram.methods import TelegramMethod
from aiogram.types import FSInputFile, InputFile
from collections import defaultdict
from datetime import datetime, timedelta
import asyncio
from typing import Dict, List, Tuple, Optional, Any
from dotenv import load_dotenv
import aiohttp
import io
import PIL
from PIL import Image
import docx
import PyPDF2
import csv
import openpyxl
import chardet
import tempfile
from PIL import Image, ExifTags
from PIL.ExifTags import TAGS
import os.path
from pathlib import Path
import shutil
from io import BytesIO
import base64
import time
import gc
from aiogram.client.default import DefaultBotProperties

# В начале файла после импортов добавим:
# Константы для авторизации
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "default_admin_password")  # Пароль админа из .env
ADMIN_IDS = set()  # Множество для хранения ID администраторов

# Множество для хранения ID пользователей с отключенными ограничениями безопасности
DISABLED_SAFETY_USERS = set()
MAX_IMAGES_PER_REQUEST = 5
MAX_IMAGES_PER_REQUEST_ADMIN = 10
IMAGE_MEMORY_LIMIT = 20
# Словарь для хранения режимов пользователей
user_modes = {}  # user_id: {"mode": "formal"|"casual", "name": str}

# Обновим константы для лимитов моделей, добавив обратно старые модели
MODEL_LIMITS = {
    "gemini-pro": {
        "free": {"rpm": 2, "tpm": 32_000, "rpd": 50},
        "paid": {"rpm": 1000, "tpm": 4_000_000, "rpd": None}
    },
    "gemini-1.5-pro-vision": {
        "free": {"rpm": 2, "tpm": 32_000, "rpd": 50},
        "paid": {"rpm": 1000, "tpm": 4_000_000, "rpd": None}
    },
    "gemini-1.5-flash": {
        "free": {"rpm": 15, "tpm": 1_000_000, "rpd": 1500},
        "paid": {"rpm": 2000, "tpm": 4_000_000, "rpd": None}
    },
    "gemini-1.5-flash-8b": {
        "free": {"rpm": 15, "tpm": 1_000_000, "rpd": 1500},
        "paid": {"rpm": 4000, "tpm": 4_000_000, "rpd": None}
    },
    "gemini-2.0-flash-exp": {
        "free": {"rpm": 5, "tpm": 50_000, "rpd": 100},
        "paid": {"rpm": 2000, "tpm": 8_000_000, "rpd": None}
    },
    "gemini-exp-1206": {
        "free": {"rpm": 3, "tpm": 40_000, "rpd": 75},
        "paid": {"rpm": 1500, "tpm": 6_000_000, "rpd": None}
    },
    "gemini-exp-1121": {
        "free": {"rpm": 3, "tpm": 40_000, "rpd": 75},
        "paid": {"rpm": 1500, "tpm": 6_000_000, "rpd": None}
    },
    "learnlm-1.5-pro-experimental": {
        "free": {"rpm": 4, "tpm": 45_000, "rpd": 80},
        "paid": {"rpm": 1800, "tpm": 7_000_000, "rpd": None}
    },
    "gemini-1.5-pro-exp-0801": {
        "free": {"rpm": 0, "tpm": 0, "rpd": 0},  # Недоступно для обычных пользователей
        "paid": {"rpm": 1000, "tpm": 4_000_000, "rpd": None}  # Доступно для админов
    },
    "gemini-1.5-flash-8b-exp-0827": {
        "free": {"rpm": 0, "tpm": 0, "rpd": 0},  # Недоступно для обычных пользователей
        "paid": {"rpm": 2000, "tpm": 4_000_000, "rpd": None}  # Доступно для админов
    }
}

# Информация о моделях
MODELS_INFO = {
    # Страница 1 - Стандартные модели
    "page1": {
    "gemini-1.5-pro": {
            "name": "Gemini Pro",
            "emoji": "🚀",
            "desc": "Самая мощная модель",
            "speed": "Средняя скорость"
        },
        "gemini-1.5-pro-vision": {
            "name": "Gemini Pro Vision",
            "emoji": "👁",
            "desc": "Продвинутая работа с изображениями",
            "speed": "Средняя скорость"
        },
        "gemini-1.5-flash": {
            "name": "Gemini Flash",
            "emoji": "⚡",
            "desc": "Быстрая модель (средние задачи)",
            "speed": "Высокая скорость"
        },
        "gemini-1.5-flash-8b": {
            "name": "Gemini Flash 8B",
            "emoji": "💨",
            "desc": "Лёгкая модель (простые задачи)",
            "speed": "Максимальная скорость"
        }
    },
    # Страница 2 - Экспериментальные модели
    "page2": {
        "gemini-2.0-flash-exp": {
            "name": "Gemini 2.0 Flash",
            "emoji": "🚀",
            "desc": "Новейшая мультимодальная модель",
            "speed": "Высокая скорость",
            "features": "Аудио, изображения, видео"
        },
        "gemini-exp-1206": {
            "name": "Gemini 1206",
            "emoji": "🎯",
            "desc": "Улучшенное качество и точность",
            "speed": "Средняя скорость",
            "features": "Улучшенный анализ"
        },
        "gemini-exp-1121": {
            "name": "Gemini 1121",
            "emoji": "💻",
            "desc": "Специализация на коде",
            "speed": "Средняя скорость",
            "features": "Код и рассуждения"
        },
        "learnlm-1.5-pro-experimental": {
            "name": "LearnLM 1.5 Pro",
            "emoji": "🧠",
            "desc": "Экспериментальная мультимодальная",
            "speed": "Высокая скорость",
            "features": "Аудио, видео, изображения"
        },
        # Добавляем новые модели
        "gemini-1.5-pro-exp-0801": {
            "name": "Близнецы 1.5 Про 0801",
            "emoji": "🧪",
            "desc": "Экспериментальная Pro версия",
            "speed": "Средняя скорость",
            "features": "Улучшенная точность"
        },
        "gemini-1.5-flash-8b-exp-0827": {
            "name": "Близнецы 1.5 Флэш-8Б 0827",
            "emoji": "💨",
            "desc": "Экспериментальная Flash версия",
            "speed": "Высокая скорость",
            "features": "Улучшенная скорость"
        }
    }
}

# Добавим словарь поддерживаемых языков
SUPPORTED_LANGUAGES = {
    "ru": "🇷🇺 Русский",
    "en": "🇬🇧 English",
    "es": "🇪🇸 Español",
    "de": "🇩🇪 Deutsch",
    "fr": "🇫🇷 Français",
    "it": "🇮🇹 Italiano",
    "zh": "🇨🇳 中文",
    "ja": "🇯🇵 日本語",
    "ko": "🇰🇷 한국어"
}

# Загружаем перем����������нные из .env файла
load_dotenv()

# Получаем токен из переменных окружения
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
if not TELEGRAM_TOKEN:
    raise ValueError("Telegram token not found in environment variables")

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("Gemini API key not found in environment variables")

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Инициализация бота и диспетчера
bot = Bot(
    token=TELEGRAM_TOKEN,
    default=DefaultBotProperties(
        parse_mode=ParseMode.MARKDOWN_V2
    )
)
dp = Dispatcher(storage=MemoryStorage())

# Создаем основной роутер
router = Router(name="main_router")

# Настройка базы данных
DB_PATH = 'bot_database.db'

def init_db():
    """Инициализация базы данных"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Создаем таблицу пользователей
        c.execute('''CREATE TABLE IF NOT EXISTS users
                     (user_id INTEGER PRIMARY KEY,
                      username TEXT,
                      first_name TEXT,
                      last_activity TIMESTAMP)''')
        
        # Создаем таблицу для сообщений
        c.execute('''CREATE TABLE IF NOT EXISTS messages
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      user_id INTEGER,
                      role TEXT,
                      content TEXT,
                      timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                      FOREIGN KEY (user_id) REFERENCES users(user_id))''')
        
        conn.commit()
        conn.close()
        logger.info("База данных успешно инициализирована")
        
    except Exception as e:
        logger.error(f"Ошибка при инициализации базы данных: {str(e)}")
        raise

def save_message(user_id: int, role: str, content: str):
    """Сохранение сообщения в БД"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Преобразуем user_id в целое число
        user_id = int(user_id)
        
        # Проверяем существование пользователя
        c.execute('SELECT 1 FROM users WHERE user_id = ?', (user_id,))
        if not c.fetchone():
            # Если пользователя нет, создаем запись
            c.execute('''INSERT INTO users (user_id, last_activity)
                        VALUES (?, ?)''', (user_id, datetime.now().isoformat()))
        
        # Сохраняем сообщение
        timestamp = datetime.now().isoformat()
        c.execute('''INSERT INTO messages (user_id, role, content, timestamp)
                     VALUES (?, ?, ?, ?)''',
                  (user_id, role, content, timestamp))

        conn.commit()
        conn.close()
        logger.info(f"Сообщение сохранено: user_id={user_id}, role={role}, timestamp={timestamp}")
        
    except Exception as e:
        logger.error(f"Ошибка при сохранении сообщения: {e}")

def get_user_history(user_id: int, limit: int = 10) -> list:
    """Получение истории сообщений пользователя"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Получаем соощения из базы данных
        c.execute('''
            SELECT role, content, timestamp 
            FROM messages 
            WHERE user_id = ? 
            ORDER BY timestamp DESC 
            LIMIT ?
        ''', (user_id, limit))
        
        messages = c.fetchall()
        conn.close()
        
        # Возвращаем в хронологическом порядке
        return list(reversed(messages))
        
    except Exception as e:
        logger.error(f"Ошибка при получении истории: {e}")
        return []

def update_user_activity(user_id: int, username: str = None, first_name: str = None):
    """Обновление информации о пользователе"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO users
                 (user_id, username, first_name, last_activity)
                 VALUES (?, ?, ?, ?)''',
              (user_id, username, first_name, datetime.now()))
    conn.commit()
    conn.close()

# Нас Gemini
genai.configure(api_key=GEMINI_API_KEY)

# Обновим конфигурации Gemini с актуальными моделями
GEMINI_CONFIGS = {
    "gemini-pro": {
        "temperature": 0.7,
        "top_p": 0.9,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-1.5-pro-vision": {
        "temperature": 0.7,
        "top_p": 0.9,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-1.5-flash": {
        "temperature": 0.9,
        "top_p": 0.95,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-1.5-flash-8b": {
        "temperature": 0.9,
        "top_p": 0.95,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-2.0-flash-exp": {
        "temperature": 0.9,
        "top_p": 0.95,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-exp-1206": {
        "temperature": 0.85,
        "top_p": 0.92,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-exp-1121": {
        "temperature": 0.85,
        "top_p": 0.92,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "learnlm-1.5-pro-experimental": {
        "temperature": 0.8,
        "top_p": 0.9,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-1.5-pro-exp-0801": {
        "temperature": 0.7,
        "top_p": 0.9,
        "top_k": 40,
        "max_output_tokens": 8192,
    },
    "gemini-1.5-flash-8b-exp-0827": {
    "temperature": 0.9,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
    }
}

# Улучшим функцию создания модели
def create_model(model_name: str) -> genai.GenerativeModel:
    """Создание модели с учетом настроек"""
    try:
        # Получаем конфигурацию для модели
        config = GEMINI_CONFIGS.get(model_name, {
            "temperature": 0.7,
            "top_p": 0.9,
            "top_k": 40,
            "max_output_tokens": 8192,
        })

        # Создаем модель с базовыми настройками безопасности
        model = genai.GenerativeModel(
            model_name=model_name,
            generation_config=config
        )
        
        return model

    except Exception as e:
        logger.error(f"Ошибка при создании модели {model_name}: {e}")
        # Возвращаем модель по умолчанию в случае ошибки
        return genai.GenerativeModel('gemini-pro')  # �����������мен��е�� значение по умолчанию

# Добвим нов��������е �����о��станты
SUPPORTED_LANGUAGES = {
    "ru": "🇷🇺 Русский",
    "en": "🇬🇧 English",
    "es": "🇪🇸 Español",
    "de": "🇩🇪 Deutsch",
    "fr": "🇫🇷 Français",
    "it": "🇮🇹 Italiano",
    "zh": "🇨🇳 中文",
    "ja": "🇯🇵 日本語",
    "ko": "🇰🇷 한국어"
}

# Добавим новые константы для работы с файлами
SUPPORTED_FILE_TYPES = {
    'image': ['.jpg', '.jpeg', '.png', '.webp', '.heic'],
    'document': ['.pdf', '.doc', '.docx', '.txt', '.rtf'],
    'spreadsheet': ['.xlsx', '.xls', '.csv'],
    'presentation': ['.ppt', '.pptx'],
    'code': ['.py', '.js', '.html', '.css', '.json', '.xml', '.yaml', '.sql'],
    'archive': ['.zip', '.rar', '.7z', '.tar', '.gz']
}

MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
MAX_IMAGE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_IMAGES_PER_GROUP = 5  # Максимальное количество изображений в группе
MEDIA_GROUP_TIMEOUT = 60  # Таймаут для группы медиа в секундах

# Улучшим функции обраоти фалв
async def process_image(image_path: str) -> tuple[PIL.Image.Image, str]:
    """Обработка зобржения с оптимизацией"""
    try:
        image = Image.open(image_path)
        
        # Конвертиуем HEIC в JPEG если нужно
        if image_path.lower().endswith('.heic'):
            new_path = image_path.rsplit('.', 1)[0] + '.jpg'
            image = image.convert('RGB')
            image.save(new_path, 'JPEG')
            image_path = new_path

        # Оптимизируем размер если нужно
        if os.path.getsize(image_path) > MAX_IMAGE_SIZE:
            while os.path.getsize(image_path) > MAX_IMAGE_SIZE:
                width, height = image.size
                image = image.resize((int(width*0.8), int(height*0.8)), Image.Resampling.LANCZOS)
                image.save(image_path, optimize=True, quality=85)

        # Извлекаем EXIF данные
        exif_data = ""
        try:
            exif = image._getexif()
            if exif:
                exif_info = []
                for tag_id in exif:
                    tag = TAGS.get(tag_id, tag_id)
                    data = exif.get(tag_id)
                    if isinstance(data, bytes):
                        data = data.decode(errors='ignore')
                    exif_info.append("{}: {}".format(tag, data))
                exif_data = "\nEXIF данные:\n" + "\n".join(exif_info)
        except:
            pass

        return image, exif_data

    except Exception as e:
        logger.error("Ошибка при обработке изображения: {}".format(e))
        raise

# Обновим константы для работы с файлами и сообщениями
MAX_MESSAGE_LENGTH = 2000  # Еще уменьшим для большей надежности
MIN_CHUNK_SIZE = 200  # Уменьшим минимальный размер части
MAX_PARTS = 8  # Ограничим количество частей
MAX_FILE_TEXT_LENGTH = 15000  # Максимальная длина текста из файла

async def extract_text_from_file(file_path: str, mime_type: str) -> str:
    """Улучшенное извлечение текста из файлов с ограничениями"""
    try:
        text_content = []
        total_length = 0
        
        if mime_type == 'application/pdf':
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                # Ограничиваем количество страниц
                max_pages = min(len(pdf_reader.pages), 10)
                
                for i in range(max_pages):
                    page_text = pdf_reader.pages[i].extract_text()
                    if total_length + len(page_text) > MAX_FILE_TEXT_LENGTH:
                        text_content.append(f"\n... (остальные {len(pdf_reader.pages) - i} страниц пропущены)")
                        break
                    text_content.append(f"[Страница {i+1}]\n{page_text}")
                    total_length += len(page_text)
                    
        elif mime_type in ['application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
            doc = docx.Document(file_path)
            
            for i, para in enumerate(doc.paragraphs):
                if total_length > MAX_FILE_TEXT_LENGTH:
                    text_content.append("\n... (остальной текст пропущен)")
                    break
                if para.text.strip():
                    text_content.append(para.text)
                    total_length += len(para.text)
                    
        elif mime_type == 'text/plain':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                lines = []
                for line in file:
                    if total_length > MAX_FILE_TEXT_LENGTH:
                        lines.append("\n... (остальной текст пропущен)")
                        break
                    lines.append(line.strip())
                    total_length += len(line)
                text_content = lines
                
        elif mime_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            wb = openpyxl.load_workbook(file_path)
            
            for sheet_name in wb.sheetnames[:3]:  # Ограничиваем количество листов
                ws = wb[sheet_name]
                sheet_content = [f"\n[Лист: {sheet_name}]"]
                
                # Ограничиваем диапазон чтения
                max_rows = min(ws.max_row, 50)
                max_cols = min(ws.max_column, 10)
                
                for row in ws.iter_rows(max_row=max_rows, max_col=max_cols):
                    row_text = " | ".join(str(cell.value or "") for cell in row)
                    if total_length > MAX_FILE_TEXT_LENGTH:
                        sheet_content.append("... (остальные данные пропущены)")
                        break
                    sheet_content.append(row_text)
                    total_length += len(row_text)
                    
                text_content.extend(sheet_content)
                
        elif mime_type == 'text/csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                csv_reader = csv.reader(file)
                rows = []
                for i, row in enumerate(csv_reader):
                    if i > 50 or total_length > MAX_FILE_TEXT_LENGTH:  # Ограничиваем количество строк
                        rows.append("... (остальные строки пропущены)")
                        break
                    row_text = " | ".join(row)
                    rows.append(row_text)
                    total_length += len(row_text)
                text_content = rows

        # Объединяем и форматируем результат
        result = "\n".join(text_content)
        if len(result) > MAX_FILE_TEXT_LENGTH:
            result = result[:MAX_FILE_TEXT_LENGTH] + "\n... (текс������ обрезан из-за большого размера)"
            
        return result.strip()
        
    except Exception as e:
        logger.error(f"Ошибка при извлечении текста: {e}")
        return f"Не удалось извлечь текст из файла: {str(e)}"

# Улучшим класс ChatSession
class ChatSession:
    def __init__(self, model_name: str = "gemini-1.5-flash", user_id: Optional[int] = None):
        self.model_name = model_name
        self.user_id = user_id
        
        # Инициализируем хранилища данных
        self.history = []
        self.long_term_memory = []
        self.message_metadata = {}  # Перемещаем инициализацию в начало
        self.message_counter = 0
        self.last_activity = datetime.now()
        
        # Оптимальные настройки для генерации
        self.generation_config = {
            "temperature": 1.0,
            "top_p": 1.0,
            "top_k": 40,
            "max_output_tokens": 8192,
            "candidate_count": 1,
        }
        
        # Настройки безопасности
        self.safety_settings = [
            {
                "category": category,
                "threshold": "BLOCK_NONE"
            }
            for category in [
                "HARM_CATEGORY_HARASSMENT",
                "HARM_CATEGORY_HATE_SPEECH",
                "HARM_CATEGORY_SEXUALLY_EXPLICIT", 
                "HARM_CATEGORY_DANGEROUS_CONTENT"
            ]
        ]
        
        # Создаем модель
        self.model = genai.GenerativeModel(
            model_name=model_name,
            generation_config=self.generation_config,
            safety_settings=self.safety_settings
        )
        
        # Инициализируем чат
        self.chat = self.model.start_chat(history=[])
        self.context_window = CHAT_SETTINGS["context_window"]
        
        # Подключаемся к БД и загружаем историю
        self.db_connection = sqlite3.connect(DB_PATH)
        self.load_history_from_db()

    def load_history_from_db(self):
        """Загрузка истории из БД"""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute('''
                SELECT role, content, timestamp 
                FROM messages 
                WHERE user_id = ? 
                ORDER BY timestamp ASC
            ''', (self.user_id,))
            
            for role, content, timestamp in cursor.fetchall():
                # Конвертируем роль в допустимый формат для Gemini
                gemini_role = "user" if role == "user" else "model"
                
                # Сохраняем сообщение
                message = {
                    "role": gemini_role,
                    "parts": [{"text": content}]
                }
                self.long_term_memory.append(message)
                
                # Сохраняем метаданные
                self.message_metadata[self.message_counter] = {
                    "timestamp": timestamp,
                    "role": role,
                    "display_role": "👤 Вы" if role == "user" else "🤖 Бот"
                }
                self.message_counter += 1
            
            # Инициализируем активную историю
            self.history = self.long_term_memory[-CHAT_SETTINGS["context_window"]:]
            
        except Exception as e:
            logger.error(f"Error loading history: {e}")
            self.history = []
            self.long_term_memory = []
            self.message_metadata = {}
            self.message_counter = 0

    async def send_message(self, text: str, stream: bool = False) -> str:
        try:
            # Создаем сообщение пользователя
            user_message = {
                "role": "user", 
                "parts": [{"text": text}]
            }
            
            # Получаем всю историю для контекста
            messages = []
            for msg in self.history:
                messages.append({
                    "role": msg["role"],
                    "parts": msg["parts"]
                })
            
            # Добавляем текущее сообщение
            messages.append(user_message)
            
            # Создаем новый чат с полной историей
            self.chat = self.model.start_chat(history=messages)
            
            # Отправляем сообщение
            if stream:
                response_text = await self.stream_response(text)
            else:
                response = await asyncio.to_thread(
                    self.chat.send_message,
                    text
                )
                response_text = response.text
            
            # Сохраняем сообщение пользователя
            self.history.append(user_message)
            self.long_term_memory.append(user_message)
            self.message_metadata[self.message_counter] = {
                "timestamp": datetime.now().isoformat(),
                "role": "user",
                "display_role": "👤 Вы"
            }
            self.message_counter += 1
            self.save_to_db("user", text)
            
            # Сохраняем ответ модели
            model_message = {
                "role": "model",
                "parts": [{"text": response_text}]
            }
            self.history.append(model_message)
            self.long_term_memory.append(model_message)
            self.message_metadata[self.message_counter] = {
                "timestamp": datetime.now().isoformat(),
                "role": "model", 
                "display_role": "🤖 Бот"
            }
            self.message_counter += 1
            self.save_to_db("model", response_text)
            
            return response_text
            
        except Exception as e:
            logger.error(f"Error in send_message: {e}")
            return "❌ Произошла ошибка при обработке запроса."

    def find_relevant_messages(self, query: str) -> List[dict]:
        """Поиск релевантных сообщений в истории"""
        relevant = []
        query_words = set(query.lower().split())
        
        for msg in self.long_term_memory:
            msg_text = msg["parts"][0]["text"].lower()
            msg_words = set(msg_text.split())
            
            # Вычисляем релевантность
            common_words = query_words.intersection(msg_words)
            relevance = len(common_words) / len(query_words) if query_words else 0
            
            if relevance >= CHAT_SETTINGS["memory_threshold"]:
                relevant.append(msg)
                
        return relevant[:CHAT_SETTINGS["max_memory_items"]]

    def save_to_db(self, role: str, content: str):
        """Сохранение сообщения в БД"""
        try:
            # Конвертируем роль assistant в model для совместимости
            db_role = "user" if role == "user" else "model"
            
            cursor = self.db_connection.cursor()
            cursor.execute('''
                INSERT INTO messages (user_id, role, content, timestamp)
                VALUES (?, ?, ?, ?)
            ''', (self.user_id, db_role, content, datetime.now().isoformat()))
            self.db_connection.commit()
        except Exception as e:
            logger.error(f"Error saving to DB: {e}")

    async def search_history(self, query: str) -> List[dict]:
        """Поиск по истории диалогов"""
        relevant = self.find_relevant_messages(query)
        results = []
        
        for i, msg in enumerate(relevant):
            metadata = self.message_metadata.get(i, {})
            results.append({
                "text": msg["parts"][0]["text"],
                "role": msg["role"],
                "timestamp": metadata.get("timestamp", ""),
                "relevance": self.calculate_relevance(query, msg["parts"][0]["text"])
            })
            
        return results

    def calculate_relevance(self, query: str, text: str) -> float:
        """Расчет релевантности текста запросу"""
        query_words = set(query.lower().split())
        text_words = set(text.lower().split())
        common_words = query_words.intersection(text_words)
        return len(common_words) / len(query_words) if query_words else 0

    def __del__(self):
        """Закрываем соединение с БД при удалении объекта"""
        if hasattr(self, 'db_connection'):
            self.db_connection.close()

    def get_chat_history(self) -> List[dict]:
        """Получение истории чата с форматированием"""
        formatted_history = []
        for msg in self.history:
            role = "👤 Вы" if msg["role"] == "user" else "🤖 Бот"
            text = msg["parts"][0]["text"]
            formatted_history.append(f"{role}: {text}")
        return formatted_history

    def clear_history(self) -> None:
        """Полная очистка истории и сброс чата"""
        self.history = []
        self.chat = self.model.start_chat()
        self.last_activity = datetime.now()

    def get_context_summary(self) -> str:
        """Получение краткого содержания контекста с форматированием"""
        if not self.history:
            return "💬 История чата пуста"
            
        last_messages = self.history[-self.context_window:]
        summary = ["📝 *Последние сообщения:*\n"]
        
        for msg in last_messages:
            role = "👤" if msg["role"] == "user" else "🤖"
            text = msg["parts"][0]["text"]
            if len(text) > 100:
                text = text[:100] + "..."
            summary.append(f"{role} {text}")
            
        return "\n".join(summary)

    def _format_response(self, text: str) -> str:
        """Простая обработка текста, позволяющая модели самой определять форматирование"""
        # Просто очищаем от лишних пробелов и переносов
        return text.strip()

    def _get_enhanced_context(self, current_text: str) -> str:
        """Формирование улучшенного контекста с учетом релевантности"""
        if not self.history:
            return ""
            
        # Получаем последние сообщения
        recent_messages = self.history[-CHAT_SETTINGS["context_window"]:]
        
        # Ищем похожие сообщения в истории
        relevant_messages = []
        for msg in recent_messages:
            if self._is_relevant(msg['parts'][0]["text"], current_text):
                relevant_messages.append(msg)
        
        # Формируем контекст
        context_parts = []
        
        # Добавляем релевантные сообщения
        if relevant_messages:
            context_parts.append("Релевантные части предыдущего разговора:")
            for msg in relevant_messages[-3:]:  # Берем последние 3 релевантных сообще��ия
                context_parts.append(msg['parts'][0]["text"])
        
        # Добавляем последние сообщения
        context_parts.append("\nПоследние сообщения:")
        for msg in recent_messages:
            context_parts.append(msg['parts'][0]["text"])
            
        return "\n".join(context_parts)

    def _is_relevant(self, history_text: str, current_text: str) -> bool:
        """Проверка релевантности сообщения текущему контексту"""
        # Простая провер��а ��а основе общих слов
        history_words = set(history_text.lower().split())
        current_words = set(current_text.lower().split())
        common_words = history_words.intersection(current_words)
        
        # Если есть до����таточно общих слов, считаем сообщение релевантным
        return len(common_words) >= 2

    def update_conversation_style(self, style: str):
        """Обновление стиля общения"""
        self.conversation_style = style
        if style == "casual":
            self.conversation_state["formality_level"] = 0.2
        else:
            self.conversation_state["formality_level"] = 0.8
        
        # Сбрасываем чат для нового контекста
        self.model = genai.GenerativeModel(self.model_name)  # Создаем новый чат

    async def _analyze_topic(self, text: str) -> str:
        """Анал����и��ет тему с��общ�����ния"""
        try:
            # Простой анализ по ключевым словам
            topics = {
                "technical": ["код", "программирование", "ошибка", "баг", "разработка"],
                "general": ["привет", "ак де������а", "погода", "нвот"],
                "help": ["помощь", "помоги", "по��скажи", "как"],
                "business": ["бизнес", "работа", "проект", "деньги"],
            }
            
            text_lower = text.lower()
            for topic, keywords in topics.items():
                if any(keyword in text_lower for keyword in keywords):
                    return topic
            return "general"
        except Exception as e:
            logger.error(f"Error in topic analysis: {e}")
            return "unknown"

    async def _analyze_sentiment(self, text: str) -> str:
        """Анализирует ��моциональный окрас сообщения"""
        try:
            # Простй анализ п ключевым словам
            positive_words = ["спасибо", "круто", "отлично", "здор��во", "класс", "хорошо"]
            negative_words = ["плохо", "ужасно", "оибка", "проблема", "���е ротает"]
            
            text_lower = text.lower()
            
            if any(word in text_lower for word in positive_words):
                return "positive"
            elif any(word in text_lower for word in negative_words):
                return "negative"
            return "neutral"

        except Exception as e:
            logger.error(f"Error in sentiment analysis: {e}")
            return "neutral"

    async def _analyze_message(self, text: str):
        """Расширенный анализ сообщеия"""
        try:
            # Анализируем тему
            topic = await self._analyze_topic(text)
            if topic != "general":
                self.conversation_state["current_topic"] = topic
                self.topics.append(topic)

            # Анализируем настроение
            sentiment = await self._analyze_sentiment(text)
            self.sentiment = sentiment
            self.conversation_state["emotional_state"] = sentiment

            # Извлекаем ключевые точки
            key_points = set(word.lower() for word in text.split() if len(word) > 3)
            self.conversation_state["key_points"].update(key_points)

            # Анализируем вопрос
            questions = [s.strip() for s in text.split('.') if '?' in s]
            self.conversation_state["questions_asked"].extend(questions)

            # Опредеяем язык
            detected_lang = self.detect_language(text)
            if detected_lang:
                self.preferred_language = detected_lang

            # Обнвл��ем глубину разговора
            self.conversation_state["conversation_depth"] += 1

            return {
                "topic": topic,
                "sentiment": sentiment,
                "key_points": key_points,
                "questions": questions,
                "language": self.preferred_language
            }

        except Exception as e:
            logger.error(f"Error in message analysis: {e}")
            return None

    def update_user_preferences(self, preferences: dict):
        """Обновляет предпочтения пользовател"""
        self.user_preferences.update(preferences)

    def get_conversation_summary(self) -> dict:
        """Возвращает сводку текущей беседы"""
        return {
            "total_messages": len(self.history),
            "main_topics": self.topics[-5:] if self.topics else [],
            "sentiment": self.sentiment,
            "language": self.preferred_language,
            "user_preferences": self.user_preferences,
            "conversation_depth": self.conversation_state["conversation_depth"],
            "key_points": list(self.conversation_state["key_points"])[:5],
            "recent_questions": self.conversation_state["questions_asked"][-3:],
            "emotional_state": self.conversation_state["emotional_state"],
            "formality_level": self.conversation_state["formality_level"]
        }

    def detect_language(self, text: str) -> str:
        """Опеделение ыка сообщения"""
        # Прота реализация определения языка
        text = text.lower()
        if any(char in 'абвгдеёжийклмнопрстуфхцчшщъыьэюя' for char in text):
            return "ru"
        elif all(ord(char) < 128 for char in text):
            return "en"
        return None

    def get_session_info(self) -> dict:
        """Получение информации о сессии"""
        return {
            "model": self.model_name,
            "language": self.preferred_language,
            "messages_count": len(self.history),
            "last_activity": self.last_activity,
            "has_image": self.last_image is not None
        }

    async def enhance_image(self, image: Image.Image) -> Image.Image:
        """Улучшение качества изображения"""
        try:
            # Базовое улучшение
            enhanced = image.copy()
            
            # Проверяем размер
            if image.size[0] * image.size[1] > 1920 * 1080:
                enhanced.thumbnail((1920, 1080), Image.Resampling.LANCZOS)
            
            # Автоматическая коррекция
            from PIL import ImageEnhance
            enhancer = ImageEnhance.Contrast(enhanced)
            enhanced = enhancer.enhance(1.2)
            enhancer = ImageEnhance.Sharpness(enhanced)
            enhanced = enhancer.enhance(1.1)
            
            return enhanced
        except Exception as e:
            logger.error(f"Ошибка при улучшении изображения: {e}")
            return image

    async def translate_text(self, text: str, target_language: str) -> str:
        """Перевод текста на другой язык"""
        try:
            # Используем API перевода
            response = await self.model.generate_content([f"Translate this to {target_language}: {text}"])
            return response.text
        except Exception as e:
            logger.error(f"Ошибка при переводе текста: {e}")
            return text

    def toggle_system_prompt(self):
        """Переключение между режимами генерации"""
        self.use_system_prompt = not self.use_system_prompt
        
        # Обновляем настройки генерации при переключении
        if self.use_system_prompt:
            # Настройки для креативного режима
            self.generation_config = {
                "temperature": 1.0,          # Максимальная креативность
                "top_p": 1.0,               # Максимальное разнообразие
                "top_k": 40,                # Больше вариантов
                "max_output_tokens": 4096,   # Длинные ответы
                "candidate_count": 1,        # Один лучший ответ
                "stop_sequences": ["User:", "Assistant:"]
            }
            
            # Промпт для креативного режима
            self.system_prompt = (
                "You are a creative and engaging AI assistant. Feel free to:\n"
                "- Use expressive language and varied tone\n"
                "- Include relevant examples and analogies\n"
                "- Break down complex topics into simple explanations\n"
                "- Add helpful context and background information\n"
                "- Use formatting to enhance readability\n"
                "- Be conversational and engaging\n"
            )
        else:
            # Настройки для точного режима
            self.generation_config = {
                "temperature": 0.4,          # Более точные ответы
                "top_p": 0.8,               # Меньше случайности
                "top_k": 20,                # Меньше вариантов
                "max_output_tokens": 2048,   # Короче ответы
                "candidate_count": 1,        # Один точный ответ
                "stop_sequences": ["User:", "Assistant:"]
            }
            
            # Промпт для точного режима
            self.system_prompt = (
                "You are a precise and concise AI assistant. Focus on:\n"
                "- Direct and accurate answers\n"
                "- Clear and factual information\n"
                "- Technical accuracy and details\n"
                "- Structured and organized responses\n"
                "- Professional and formal tone\n"
            )
        
        return self.use_system_prompt

# После других классов добавим класс APIUsageTracker
class APIUsageTracker:
    def __init__(self):
        self.requests_per_minute = defaultdict(list)
        self.requests_per_day = defaultdict(list)
        self.current_model = "gemini-pro"  # модель по умолчанию
        self.is_paid = False
        self.user_stats = defaultdict(lambda: {
            "total_requests": 0,
            "total_tokens": 0,
            "favorite_model": None,
            "last_request": None
        })

    # Доавляем метод can_make_request
    async def can_make_request(self, user_id: str) -> Tuple[bool, str]:
        """Проверка возможности выполнения запроса"""
        # Если польователь админ - пропускаем проверки
        if user_id in ADMIN_IDS:
            return True, ""
        
        self.clean_old_requests(user_id)
        
        # Получаем лимиты текущей модели
        limits = MODEL_LIMITS[self.current_model]["paid" if self.is_paid else "free"]
        rpm_count = len(self.requests_per_minute[user_id])
        rpd_count = len(self.requests_per_day[user_id])

        if rpm_count >= limits["rpm"]:
            wait_time = self._get_wait_time(user_id, "minute")
            message = escape_markdown_v2(f"⚠️ Достигнут лимит запросов в минуту. Подождите {wait_time} сек.")
            return False, message

        if limits["rpd"] and rpd_count >= limits["rpd"]:
            wait_time = self._get_wait_time(user_id, "day")
            message = escape_markdown_v2(f"⚠️ Достигнут дневной лимит. Новые запросы будут доступны через {wait_time} мин.")
            return False, message

        return True, ""

    def _get_wait_time(self, user_id: str, period: str) -> int:
        """Получение врмни ожидания до следущего запроса"""
        current_time = datetime.now()
        if period == "minute":
            oldest_request = min(self.requests_per_minute[user_id])
            return 60 - (current_time - oldest_request).seconds
        else:  # day
            oldest_request = min(self.requests_per_day[user_id])
            return (24 * 60) - ((current_time - oldest_request).seconds // 60)

    def track_request(self, user_id: str):
        """Отслеживание запроса пользователя"""
        current_time = datetime.now()
        
        # Добавляем запрос в статистику минуты
        self.requests_per_minute[user_id].append(current_time)
        
        # Добавляем запрос в статистику дня
        self.requests_per_day[user_id].append(current_time)
        
        # Обновляем статистику пользователя
        stats = self.user_stats[user_id]
        stats["total_requests"] += 1
        stats["last_request"] = current_time
        
        # Очищаем старые запросы
        self.clean_old_requests(user_id)

    def clean_old_requests(self, user_id: str):
        """Очистка старых запросов"""
        current_time = datetime.now()
        minute_ago = current_time - timedelta(minutes=1)
        day_ago = current_time - timedelta(days=1)
        
        # Оищаем зпросы старше минуты
        self.requests_per_minute[user_id] = [
            t for t in self.requests_per_minute[user_id]
            if t > minute_ago
        ]
        
        # Очищаем запросы сташе дня
        self.requests_per_day[user_id] = [
            t for t in self.requests_per_day[user_id]
            if t > day_ago
        ]

    def get_usage_stats(self, user_id: str) -> dict:
        """Получение статистии испоьзования"""
        self.clean_old_requests(user_id)
        
        stats = self.user_stats[user_id]
        current_limits = MODEL_LIMITS[self.current_model]["paid" if self.is_paid else "free"]
        
        return {
            "total_requests": stats["total_requests"],
            "total_tokens": stats["total_tokens"],
            "requests_today": len(self.requests_per_day[user_id]),
            "requests_minute": len(self.requests_per_minute[user_id]),
            "favorite_model": stats["favorite_model"] or self.current_model,
            "current_model": self.current_model,
            "is_paid": self.is_paid,
            "last_update": datetime.now(),
            "limits": current_limits
        }

# Создаем глобальный экземпляр трекера
api_tracker = APIUsageTracker()

# оздаем глобальные объкты
chat_sessions = {}

def get_chat_keyboard():
    builder = InlineKeyboardBuilder()
    builder.add(InlineKeyboardButton(text="Новый диалог", callback_data="new_chat"))
    builder.add(InlineKeyboardButton(text="История чта", callback_data="show_history"))
    builder.add(InlineKeyboardButton(text="Экспорт истории", callback_data="export_history"))
    builder.add(InlineKeyboardButton(text="омощь", callback_data="help"))
    return builder.as_markup()

def setup_bot_commands():
    """Установка команд бота"""
    commands = [
        types.BotCommand(command="start", description="🚀 Запустить бота"),
        types.BotCommand(command="help", description="❓ Помощь и инструкции"),
        types.BotCommand(command="new", description="🔄 Новый диалог"),
        types.BotCommand(command="history", description="📜 История диалога"),
        types.BotCommand(command="export", description="📤 Экспорт истории в файл"),
        types.BotCommand(command="model", description="🤖 Выбор модели AI"),
        types.BotCommand(command="stats", description="📊 Статистика использования"),
        types.BotCommand(command="mode", description="⚙️ Режим общения"),
        types.BotCommand(command="search", description="🔍 Поиск по истории"),
        types.BotCommand(command="summary", description="📝 Сводка диалога"),
        types.BotCommand(command="clear", description="🗑 Очистить историю"),
        types.BotCommand(command="settings", description="⚙️ Настройки"),
        types.BotCommand(command="context", description="🧠 Текущий контекст"),
        types.BotCommand(command="analyze", description="🔬 Анализ изображений"),
        types.BotCommand(command="auth", description="🔑 Авторизация админа"),
    ]
    return commands

# Сначала все обработчики команд
@router.message(Command("start"), flags={"command": True})
async def start_handler(message: types.Message):
    """Обработчик команды /start"""
    try:
        user_id = str(message.from_user.id)
        
        # Создаем новую сессию при старте
        chat_sessions[user_id] = ChatSession(api_tracker.current_model)
        
        # Обновляем информацию о пользовате��е
        update_user_activity(
            user_id=message.from_user.id,
            username=message.from_user.username,
            first_name=message.from_user.first_name
        )
        
        # Содаем клавиатуру
        keyboard = InlineKeyboardBuilder()
        keyboard.row(
            InlineKeyboardButton(
                text="📤 История чата",
                callback_data="show_history"
            ),
            InlineKeyboardButton(
                text=" Помощь",
                callback_data="help"
            )
        )
        keyboard.row(
            InlineKeyboardButton(
                text="🔄 новый диалог",
                callback_data="new_chat"
            ),
            InlineKeyboardButton(
                text="📊 Статистика",
                callback_data="show_stats"
            )
        )
        
        welcome_text = (
            f"*Доб��о по��аловать, {message.from_user.first_name}!*\n\n"
            "🤖 * ваш AI-помощник на базе Gemini*\n\n"
            "*Мои во��можности:*\n"
            "• Ответы на вопросы\n"
            "• Анализ и��ображений\n"
            "• Работа с документами\n"
            "• Написание текстов\n"
            "• Помощ с кодом\n\n"
            "*Основные кома��д��:*\n"
            "🔄 /new - Нов��й диало����\n"
            "🔧 /model - Выбор модели\n"
            "📊 /stats - Статистика\n"
            "❓ /help - Помощь\n\n"
            "_Отправьте сообщени или изображение, тобы начат общние_"
        )
        
        await message.reply(
            welcome_text,
            parse_mode="Markdown",
            reply_markup=keyboard.as_markup()
        )
        
        # Оправляем иформацию о текущей модели
        model_info = {
            "gemini-1.5-pro": {"name": "Gemini Pro", "emoji": "������"},
            "gemini-1.5-flash": {"name": "Gemini Flash", "emoji": "⚡"},
            "gemini-1.5-flash-8b": {"name": "Gemini Flash 8B", "emoji": "💨"}
        }.get(api_tracker.current_model, {"name": api_tracker.current_model, "emoji": "🤖"})
        
        model_text = (
            f"{model_info['emoji']} *Текущая модель: {model_info['name']}*\n\n"
            "Используйте /model для смены модели"
        )
        
        await message.answer(model_text, parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Ошиба в start_handler: {e}")
        await message.reply(
            "��� Пр��изошла ошибка при запуске. Попробуйте еще раз.",
            parse_mode="Markdown"
        )

# Добавляем обработчик auth сразу после start
@router.message(Command("auth"))
async def auth_handler(message: types.Message):
    """Обработчик команды авторизации администратора"""
    try:
        # Получаем пароль из сообщения
        parts = message.text.split(maxsplit=1)
        if len(parts) != 2:
            await message.reply(
                "❌ *Неверный формат команды*\n"
                "Используйте: /auth <пароль>",
                parse_mode="Markdown"
            )
            return

        password = parts[1].strip()
        user_id = str(message.from_user.id)
        
        # Проверяем пароль
        if password == ADMIN_PASSWORD:
            ADMIN_IDS.add(user_id)
            api_tracker.is_paid = True
            
            if user_id in chat_sessions:
                chat_sessions[user_id] = ChatSession(api_tracker.current_model)
            
            await message.reply(
                "🎉 *Авторизация успешна!*\n\n"
                "✅ Вы получили права администратора\n"
                "✅ Лимиты на запросы сняты\n"
                "✅ Доступны все модели без ограничений",
                parse_mode="Markdown"
            )
            
            await message.delete()
            logger.info(f"Пользователь {user_id} получил права администратора")
            
        else:
            await message.reply(
                "❌ *Неверный пароль!*\n\n"
                "Доступ запрещен.",
                parse_mode="Markdown"
            )
            logger.warning(f"Неудачная попытка авторизации от пользователя {user_id}")
            
    except Exception as e:
        logger.error(f"Ошибка в auth_handler: {e}")
        await message.reply(
            "😔 *Произошла ошибка при авторизации*",
            parse_mode="Markdown"
        )

@router.message(Command("help"))
async def help_handler(message: types.Message):
    help_text = escape_markdown_v2(
        "🤖 Я ваш AI-помощник на базе Gemini\n\n"
        "Мои возможности:\n\n"
        "🔹 Ответы на вопросы:\n"
        "Я отвечу на ваши вопросы по любым темам. Спрашивайте!\n\n"
        "🔹 Генерация текста:\n"
        "Я могу писать тексты на разных языках: статьи, истории, стихи.\n\n"
        "🔹 Перевод:\n"
        "Помогу с переводом между разными языками.\n\n"
        "🔹 Анализ текста:\n"
        "Могу обобщать и анализировать тексты.\n\n"
        "🔹 Помощь с кодом:\n"
        "Помогу с программированием и отладкой.\n\n"
        "Доступные команды:\n"
        "/start - Начать общение\n"
        "/new - Новый диалог\n"
        "/history - История чата\n"
        "/export - Экспорт истории\n"
        "/model - Выбор модели\n"
        "/stats - Статистика\n"
        "/help - Это сообщение\n\n"
        "Совет: Чем точнее вопрос, тем полезнее будет ответ."
    )
    await message.reply(
        text=help_text,
        parse_mode=ParseMode.MARKDOWN_V2
    )

@router.message(Command("model"), flags={"command": True})
async def model_handler(message: types.Message):
    """Обработчик команды выбора модели"""
    try:
        # Создаем клавиатуру для первой страницы
        keyboard = InlineKeyboardBuilder()
        
        # Добавяем кнопки моделей первой страницы
        for model_id, info in MODELS_INFO["page1"].items():
            keyboard.add(InlineKeyboardButton(
                text=f"{info['emoji']} {info['name']}",
                callback_data=f"model_{model_id}"
            ))
        
        # Добавляем кнопку перехода на вторую страниц
            keyboard.add(InlineKeyboardButton(
            text="➡️ Экспериментальные модели",
            callback_data="model_page_2"
        ))
        
        # Рас����оагаем копки в два столбц
        keyboard.adjust(2)
        
        # Формируем текс сообщния
        model_text = (
            "🤖 *Вы��ерите модель:*\n\n"
            "*Стандартные модели:*\n\n"
        )
        
        for info in MODELS_INFO["page1"].values():
            model_text += (
                f"{info['emoji']} *{info['name']}*\n"
                f"└ {info['desc']}\n"
                f"└ Сорость: {info['speed']}\n\n"
            )

        await message.reply(
            model_text,
            reply_markup=keyboard.as_markup(),
            parse_mode="Markdown"
        )

    except Exception as e:
        logger.error(f"Ошибка в model_handler: {e}")
        await message.reply("Произошла ошибка при отображении моделей")

@router.message(Command("stats"), flags={"command": True})
async def stats_handler(message: types.Message):
    """Обработчик команды статистики"""
    try:
        user_id = str(message.from_user.id)
        stats = api_tracker.get_usage_stats(user_id)
        
        stats_text = (
            "📊 *Статистика использования:*\n\n"
            f"🤖 Текущая одел: {stats['current_model']}\n"
            f"��� Тип доступа: {'Premium' if stats['is_paid'] else 'Базовый'}\n\n"
            f"📈 *Запросы:*\n"
            f"• Всего: {stats['total_requests']}\n"
            f"• За сегодня: {stats['requests_today']}\n"
            f"• За минуту: {stats['requests_minute']}\n\n"
            f"🎯 *Лимиты:*\n"
            f"• Запросов в минуту: {stats['limits']['rpm']}\n"
            f"• Токенов в минуту: {stats['limits']['tpm']:,}\n"
            f"• Запросов в день: {stats['limits']['rpd'] or 'Без ограничений'}\n\n"
            f"⏰ Обновлено: {stats['last_update'].strftime('%H:%M:%S')}"
        )
        
        await message.reply(stats_text, parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Ошибка в stats_handler: {e}")
        await message.reply("Произошла ошибка при получнии статстики")

@router.message(Command("new"))
async def new_chat_handler(message: types.Message):
    """Начать новый диалог"""
    try:
        user_id = str(message.from_user.id)
        
        # Создаем новую сессию чата
        chat_sessions[user_id] = ChatSession(api_tracker.current_model)
        
        await message.reply(
            "🔄 *Начат новый диалог!*",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в new_chat_handler: {e}")
        await message.reply(
            "❌ *Произошла ошибка при создании нового диалога*",
            parse_mode="Markdown"
        )

@router.message(F.photo)
async def handle_photo(message: Message):
    """Обработчик фотографий"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем лимиты запросов
        can_request, error_message = await api_tracker.can_make_request(user_id)
        if not can_request:
            await message.reply(error_message)
            return

        # Получаем фото максимального размера
        photo = message.photo[-1]
        
        # Создаем временную директорию для пользовател
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Загружаем файл
            file_info = await bot.get_file(photo.file_id)
            file_path = os.path.join(temp_dir, f"image_{photo.file_id}.jpg")
            await bot.download_file(file_info.file_path, file_path)
            
            # Обабатываем изображение
            with Image.open(file_path) as img:
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                processed_image = img.copy()

            # Создаем или получаем сессию с моделью
            if user_id not in chat_sessions:
                chat_sessions[user_id] = ChatSession("gemini-1.5-pro-vision", user_id)
            session = chat_sessions[user_id]

            # Формируем промпт
            prompt = message.caption if message.caption else "Опиши что изображено на этой фотографии"

            # Используем модель из сессии пользователя
            model = create_model(session.model_name)
            
            # Отправляем запрос к модели
            response = model.generate_content([
                prompt,
                processed_image
            ])

            # Проверям и отправляем ответ
            if response and hasattr(response, 'text'):
                # Сохраняем историю
                save_message(user_id, "user", f"[Изобажение] {prompt}")
                save_message(user_id, "assistant", response.text)
                
                # Отслеживаем запрос
                api_tracker.track_request(user_id)
                
                # Отправляем ответ
                await send_long_message(message, response.text)
            else:
                await message.reply("Не удалось полчить ответ от модели")

        except Exception as e:
            logger.error(f"Ошибка при бработке изображения: {e}")
            await message.reply(
                "Произошла ошка при бработке изображения. "
                "Возожно, текущая модель не поддерживает работу с изображениями. "
                "Используйте /model для выбора модели с поддержкой изображений."
            )
            
        finally:
            # Очищаем временные файлы
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                
    except Exception as e:
        logger.error(f"Критическая ошибка в handle_photo: {e}")
        await message.reply("Произошла критическая ошибка при обработке фотографии")

# Добавим функцию для преобразования изображения в байты
def image_to_bytes(image: Image.Image) -> bytes:
    """Преобразование изображения в байты"""
    try:
        with io.BytesIO() as bio:
            image.save(bio, format='JPEG')
            return bio.getvalue()
    except Exception as e:
        logger.error(f"Ошибка при конвертации изображения в байты: {e}")
        raise

# Добавим функцию для безопасного экранирования Markdown
def escape_markdown(text: str) -> str:
    """Экранирование специальных символов Markdown"""
    escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    return ''.join(f'\\{c}' if c in escape_chars else c for c in text)

@router.message(F.media_group_id)
async def handle_media_group(message: Message):
    """Обаботчик группы фотографий"""
    try:
        user_id = str(message.from_user.id)
        media_group_id = message.media_group_id

        # Проверяем, есть ли фото в сообщении
        if not message.photo:
            return

        # Создаем или получаем сессию
        if user_id not in chat_sessions:
            chat_sessions[user_id] = ChatSession("gemini-pro-vision", user_id)

        # Инициализируем группу медиа если её еще нет
        if media_group_id not in media_groups:
            media_groups[media_group_id] = {
                "images": [],
                "user_id": user_id,
                "timestamp": datetime.utcnow(),
                "waiting_prompt": True
            }

        # Получаем фото максимального размера
        photo = message.photo[-1]
        
        # Создаем временную директорию
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Загружаем и обрабатываем фото
            file_info = await bot.get_file(photo.file_id)
            file_path = os.path.join(temp_dir, f"image_{photo.file_id}.jpg")
            await bot.download_file(file_info.file_path, file_path)
            
            with Image.open(file_path) as img:
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                processed_image = img.copy()
            
            media_groups[media_group_id]["images"].append(processed_image)

            # Если это последнее фото в группе
            if len(message.media_group) == len(media_groups[media_group_id]["images"]):
                await message.reply(
                    "✅ Группа фотографий получена!\n"
                    "Отправьте ваш вопрос или описание.",
                    parse_mode="Markdown"
                )
                
        except Exception as e:
            logger.error(f"Ошибка при обработке фото в группе: {e}")
            await message.reply("Произошла ошибка при обработке фотографии в группе")
        finally:
            # Оищаем временные файлы
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

    except Exception as e:
        logger.error(f"Ошибка в handle_media_group: {e}")
        await message.reply(
            "Произошла ошибка при обработке группы фотографий. "
            "Попробуйте отправить меньше фотографий или повторите позже."
        )

@router.message(F.document)
async def handle_document(message: Message):
    """Улучшенный обработчик документов"""
    try:
        user_id = str(message.from_user.id)
        
        can_request, error_message = await api_tracker.can_make_request(user_id)
        if not can_request:
            await send_formatted_message(message, error_message, "warning")
            return

        temp_dir = tempfile.mkdtemp()
        try:
            # Скачиваем файл
            file_info = await bot.get_file(message.document.file_id)
            file_path = os.path.join(temp_dir, message.document.file_name)
            await bot.download_file(file_info.file_path, file_path)
            
            # Обрабатываем файл
            result = await process_file(message, file_path, message.document.mime_type)
            
            # Отправляем результат
            await send_formatted_message(message, result)
            
        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                
    except Exception as e:
        logger.error(f"Ошибка при обработке документа: {e}")
        await send_formatted_message(message, str(e), "error")

@router.message(F.text & ~F.text.startswith('/'))  # Обрабатвать только ткстовые сообщения, не начинающиеся с '/'
async def message_handler(message: Message) -> None:
    """Улучшенный обработчик текстовых сообщений"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем лимиты запросов
        can_request, error_message = await api_tracker.can_make_request(user_id)
        if not can_request:
            await message.answer(
                text=error_message,
                parse_mode=ParseMode.MARKDOWN_V2
            )
            return

        # Получаем или создаем сессию
        if user_id not in chat_sessions:
            chat_sessions[user_id] = ChatSession(api_tracker.current_model, user_id)
        session = chat_sessions[user_id]

        # Используем ChatActionSender для отображения набора текста
        async with ChatActionSender.typing(bot=message.bot, chat_id=message.chat.id):
            try:
                # Получаем ответ от модели
                response = await session.send_message(message.text)
                
                # Проверяем на сообщения безопасности
                if response.startswith("⚠️"):
                    await message.answer(
                        text=escape_markdown_v2(response),
                        parse_mode=ParseMode.MARKDOWN_V2
                    )
                    return
                
                # Сохраняем сообщения
                save_message(user_id, "user", message.text)
                save_message(user_id, "assistant", response)
                
                # Отслеживаем использование
                api_tracker.track_request(user_id)
                
                # Отправляем ответ
                await send_long_message(message, response)
                
            except Exception as model_error:
                logger.error(f"Model error: {model_error}")
                error_text = escape_markdown_v2(
                    "❌ Произошла ошибка при получении ответа. "
                    "Попробуйте переформулировать запрос."
                )
                await message.answer(
                    text=error_text,
                    parse_mode=ParseMode.MARKDOWN_V2
                )

    except Exception as e:
        logger.error(f"Error in message_handler: {e}")
        error_text = escape_markdown_v2(
            "❌ Произошла ошибка при обработке сообщения. "
            "Попробуйте еще раз или начните новый диалог командой /new"
        )
        await message.answer(
            text=error_text,
            parse_mode=ParseMode.MARKDOWN_V2
        )

@router.message(F.text, flags={"long_operation": "typing"})
async def message_handler(message: types.Message):
    """Обработчик текстовых сообщений"""
    try:
        user_id = str(message.from_user.id)
        
        # Создаем сессию если её нет
        if user_id not in chat_sessions:
            chat_sessions[user_id] = ChatSession(api_tracker.current_model, user_id)
        
        session = chat_sessions[user_id]
        
        # Проверяем специальные команды
        if message.text.lower() == "покажи историю":
            summary = session.get_context_summary()
            await message.reply(f"📝 История диалога:\n\n{summary}")
            return
            
        # Получаем ответ от модели
        response = await session.send_message(message.text)
        
        # Отправляем ответ
        await send_long_message(message, response)
        
    except Exception as e:
        logger.error(f"шибка в message_handler: {e}")
        await message.reply(
            "Произошла ошибка при обработке сообщения. Попробуйте позже или начните новый диалог командой /clear"
        )

# Обработчик callback запроов должен быть до main()
@router.callback_query()
async def callback_handler(callback: types.CallbackQuery):
    """Обработчик callback запросов"""
    try:
        user_id = str(callback.from_user.id)
        is_admin = user_id in ADMIN_IDS

        if callback.data == "model_page_2":
            # Показываем вторую страницу с эксперименталными моделями
            keyboard = InlineKeyboardBuilder()
            
            for model_id, info in MODELS_INFO["page2"].items():
                keyboard.add(InlineKeyboardButton(
                    text=f"{info['emoji']} {info['name']}",
                    callback_data=f"model_{model_id}"
                ))
            
            keyboard.add(InlineKeyboardButton(
                text="⬅️ Наа к стандартным моделям",
                callback_data="model_page_1"
            ))
            
            keyboard.adjust(2)
            
            model_text = "🧪 *Экспериментальные модели:*\n\n"
            
            for info in MODELS_INFO["page2"].values():
                model_text += (
                    f"{info['emoji']} *{info['name']}*\n"
                    f"└ {info['desc']}\n"
                    f"└ Скорость: {info['speed']}\n"
                    f"└ Возможноти: {info.get('features', 'Стандартные')}\n\n"
                )

            await callback.message.edit_text(
                model_text,
                reply_markup=keyboard.as_markup(),
                parse_mode="Markdown"
            )

        elif callback.data == "model_page_1":
            keyboard = InlineKeyboardBuilder()
            
            for model_id, info in MODELS_INFO["page1"].items():
                keyboard.add(InlineKeyboardButton(
                    text=f"{info['emoji']} {info['name']}",
                    callback_data=f"model_{model_id}"
                ))
            
            keyboard.add(InlineKeyboardButton(
                text="➡️ Экспериметальные модели",
                callback_data="model_page_2"
            ))
            
            keyboard.adjust(2)
            
            model_text = (
                "🤖 *Выберите модель:*\n\n"
                "*Стандартные модели:*\n\n"
            )
            
            for info in MODELS_INFO["page1"].values():
                model_text += (
                    f"{info['emoji']} *{info['name']}*\n"
                    f"└ {info['desc']}\n"
                    f"└ Скорость: {info['speed']}\n\n"
                )

            await callback.message.edit_text(
                model_text,
                reply_markup=keyboard.as_markup(),
                parse_mode="Markdown"
            )

        elif callback.data.startswith("model_"):
            model_name = callback.data.replace("model_", "")
            
            # Поверям дотупность модели
            if not is_admin and MODEL_LIMITS[model_name]["free"]["rpm"] == 0:
                await callback.answer(
                    "Эта модеь дступна только администраторам", 
                    show_alert=True
                )
                return
            
            # Создаем нвую сессию с выбранной моделью
            chat_sessions[user_id] = ChatSession(model_name, user_id)
            api_tracker.current_model = model_name
            
            # Ищем информцию о модели
            model_info = None
            for page in MODELS_INFO.values():
                if model_name in page:
                    model_info = page[model_name]
                    break
            
            if model_info:
                await callback.message.edit_text(
                    f"{model_info['emoji']} Модел изменена на *{model_info['name']}*\n"
                    "Можете начать новый диалог!",
                    parse_mode="Markdown"
                )
                await callback.answer(
                    f"Модель изменена на {model_info['name']}!"
                )
            else:
                await callback.answer(
                    "Ошибка при выборе модели", 
                    show_alert=True
                )

        elif callback.data.startswith("mode_"):
            mode = callback.data.split("_")[1]
            
            # Обновлем режим пользователя
            if user_id not in user_modes:
                user_modes[user_id] = {}
            user_modes[user_id]["mode"] = mode
            
            # Создаем новую сессию с обновленными настройками
            if user_id in chat_sessions:
                session = chat_sessions[user_id]
                session.update_conversation_style(mode)
            
            # Отправляем подтверждение
            if mode == "casual":
                await callback.message.edit_text(
                    "😊 Отлично! Теперь общаемся по-дружески. Как дела?",
                    parse_mode="Markdown"
                )
            else:
                await callback.message.edit_text(
                    "👔 Перехожу на формальный стиль общения. Чем могу помочь?",
                    parse_mode="Markdown"
                )
            
            await callback.answer("Режи общения измнен!")
            
        elif callback.data == "new_chat":
            chat_sessions[user_id] = ChatSession(api_tracker.current_model)
            await callback.answer("Начат новый диалог!")
            await callback.message.edit_text("🔄 Начат новый диалог!")
            
        elif callback.data == "show_history":
            await cmd_history(callback.message)
            await callback.answer()
            
        elif callback.data == "export_history":
            await cmd_export(callback.message)
            await callback.answer()
            
        elif callback.data == "help":
            await help_handler(callback.message)
            await callback.answer()
            
        elif callback.data == "clear_history":
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('DELETE FROM messages WHERE user_id = ?', (callback.from_user.id,))
            conn.commit()
            conn.close()
            
            await callback.answer("История очищена!")
            await callback.message.edit_text("🗑 история чата очищена!")
        
        elif callback.data == "refresh_stats":
            # Обновляем статистику
            await stats_handler(callback.message)
            await callback.answer("Статистика обновлена!")
            
        elif callback.data == "detailed_stats":
            user_id = str(callback.from_user.id)
            stats = api_tracker.get_usage_stats(user_id)
            
            # Получаем исторю использования из базы данных
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('''
                SELECT DATE(timestamp), COUNT(*) 
                FROM messages 
                WHERE user_id = ? 
                GROUP BY DATE(timestamp)
                ORDER BY DATE(timestamp) DESC
                LIMIT 7
            ''', (callback.from_user.id,))
            daily_stats = c.fetchall()
            conn.close()
            
            detailed_text = "* Подробная статистика*\n\n"
            
            # Доавлем статистику по дням
            detailed_text += "*Использование по дням:*\n"
            for date, count in daily_stats:
                detailed_text += f" {date}: {count} сообщений\n"
            
            await callback.message.reply(
                detailed_text,
                parse_mode="Markdown"
            )
            await callback.answer()
        
        elif callback.data == "refresh_history":
            await cmd_history(callback.message)
            await callback.answer("История обновлена!")
        
        elif callback.data == "toggle_safety":
            user_id = str(callback.from_user.id)
            
            if user_id not in ADMIN_IDS:
                await callback.answer("Доступ запрещен!", show_alert=True)
                return
            
            if user_id in DISABLED_SAFETY_USERS:
                DISABLED_SAFETY_USERS.remove(user_id)
                is_disabled = False
                message = "🔓 Ограничения безопасноси вкючены"
            else:
                DISABLED_SAFETY_USERS.add(user_id)
                is_disabled = True
                message = "🔒 Ограничения безопасности отключены"
            
            # Пересоздем сессю с новыми нстройками
            if user_id in chat_sessions:
                chat_sessions[user_id] = ChatSession(api_tracker.current_model)
            
            # Оновляем сообщение
            status_text = (
                "🛡 *Управление ограничениями безопасности*\n\n"
                f"Текущий статус: {'🔓 Отключены' if is_disabled else '🔒 Включены'}\n\n"
                "⚠️ *Внимание:* Отключение ограничний может привести к генераии "
                "нежелательного контента. Используйте с осторожностью!"
            )
            
            keyboard = InlineKeyboardBuilder()
            keyboard.add(InlineKeyboardButton(
                text="🔓 Отключить ограничения" if not is_disabled else "🔒 Включить ограничения",
                callback_data="toggle_safety"
            ))
            
            await callback.message.edit_text(
                status_text,
                reply_markup=keyboard.as_markup(),
                parse_mode="Markdown"
            )
            
            await callback.answer(message, show_alert=True)
        
    except Exception as e:
        logger.error(f"Ошибка в callback_handler: {e}")
        await callback.answer(
            "Произошла ошибка при обраотке зпроса!", 
            show_alert=True
        )

@router.message(Command("export"), flags={"command": True})
async def cmd_export(message: types.Message):
    logger.info("Вызван обработчик команды /export")
    """Обработчик команды /export"""
    try:
        user_id = message.from_user.id
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
            SELECT role, content, timestamp 
            FROM messages 
            WHERE user_id = ? 
            ORDER BY timestamp ASC
        ''', (user_id,))
        
        messages = c.fetchall()
        conn.close()

        if not messages:
            await message.reply("История чата пуста")
            return

        export_data = {
            "user_id": user_id,
            "username": message.from_user.username,
            "export_date": datetime.now().isoformat(),
            "messages": [
                {
                    "role": msg[0],
                    "content": msg[1],
                    "timestamp": msg[2]
                }
                for msg in messages
            ]
        }

        filename = f"chat_history_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            await message.reply_document(
                FSInputFile(filename),
                caption="📤 Эспорт истории чата"
            )
        finally:
            if os.path.exists(filename):
                os.remove(filename)
                
    except Exception as e:
        logger.error(f"Ошибка в cmd_export: {e}")
        await message.reply("Произошла ошибка при экспорте истории")

@router.message(Command("history"), flags={"command": True})
async def history_handler(message: types.Message):
    """Показать историю чата"""
    try:
        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            await message.reply("У вас нет активной сессии чата")
            return
            
        session = chat_sessions[user_id]
        summary = session.get_context_summary()
        
        await message.reply(
            f"📝 *История диалога:*\n\n{summary}",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в history_handler: {e}")
        await message.reply("Произошла ошибка при получении истории")

@router.message(Command("mode"), flags={"command": True})
async def mode_handler(message: types.Message):
    logger.info("Вызван обработчик команды /mode")
    """Обработчик команды /mode"""
    user_id = str(message.from_user.id)
    
    # Создаем лавиатуру для выбора реима
    keyboard = InlineKeyboardBuilder()
    keyboard.add(InlineKeyboardButton(
        text="👔 Формальный",
        callback_data="mode_formal"
    ))
    keyboard.add(InlineKeyboardButton(
        text="👋 Дружеси",
        callback_data="mode_casual"
    ))
    
    current_mode = user_modes.get(user_id, {}).get("mode", "formal")
    mode_text = "формалном" if current_mode == "formal" else "друеском"
    
    await message.reply(
        f"🔄 *Выберите режим общения*\n\n"
        f"Сейчас я обаюсь  *{mode_text}* режиме.\n\n"
        " *Формальный* - делово стль, четкие ответы\n"
        "👋 *Дружеский* - нефомальное общение, эмодзи, шутки",
        reply_markup=keyboard.as_markup(),
        parse_mode="Markdown"
    )

@router.message(Command("summary"), flags={"command": True})
async def summary_handler(message: types.Message):
    """Показывает сводку текущей беседы"""
    try:
        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            await message.reply("Нет активной сссии чата.")
            return

        session = chat_sessions[user_id]
        summary = session.get_conversation_summary()

        summary_text = (
            "📊 *Сводка беседы:*\n\n"
            f" Всего сообщений: {summary['total_messages']}\n"
            f"🎯 Основные темы: {', '.join(summary['main_topics']) if summary['main_topics'] else 'Не определены'}\n"
            f"😊 Тон бесеы: {summary['sentiment']}\n"
            f"🌐 Язык: {SUPPORTED_LANGUAGES.get(summary['language'], 'Не определен')}\n\n"
            "*Предпотения пользователя:*\n"
        )

        for key, value in summary['user_preferences'].items():
            summary_text += f"• {key}: {value}\n"

        await message.reply(summary_text, parse_mode="Markdown")

    except Exception as e:
        logger.error(f"Error in summary_handler: {e}")
        await message.reply("Произошла ошибка при получении сводки.")

@router.message(Command("clear"))
async def clear_handler(message: types.Message):
    """Очистка истории чата"""
    try:
        user_id = str(message.from_user.id)
        
        # Создаем новую сессию чата
        chat_sessions[user_id] = ChatSession(api_tracker.current_model)
        
        # Очищаем историю в базе данных
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('DELETE FROM messages WHERE user_id = ?', (message.from_user.id,))
        conn.commit()
        conn.close()
        
        await message.reply(
            "🗑 *История чата очищена*\n"
            "Можете начать новый диалог!",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в clear_handler: {e}")
        await message.reply(
            "❌ *Произошла ошибка при очистке истории*",
            parse_mode="Markdown"
        )

@router.message(Command("context"), flags={"command": True})
async def context_handler(message: types.Message):
    """Показывает текущий контекст разговора"""
    try:
        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            await message.reply("Нет активной сессии чата.")
            return

        session = chat_sessions[user_id]
        state = session.conversation_state

        context_text = (
            "🧠 *Текущий контекст разговора:*\n\n"
            f"📌 Текущая тема: {state['current_topic'] or 'Не определена'}\n"
            f"📊 Глубина разговора: {state['conversation_depth']}\n\n"
            "*Ключевы моменты:*\n"
            f"{', '.join(state['key_points'][:5]) or 'Нт'}\n\n"
            "*Последние вопросы:*\n"
            f"{', '.join(state['questions_asked'][-3:]) or 'Нет'}\n\n"
            "*Предлагамые темы:*\n"
            f"{', '.join(state['follow_up_suggestions']) or 'Нет'}"
        )

        await message.reply(context_text, parse_mode="Markdown")

    except Exception as e:
        logger.error(f"Error in context_handler: {e}")
        await message.reply("Произошла ошибка при олучении контекста.")

@router.message(Command("safety"), flags={"command": True})
async def safety_handler(message: types.Message):
    """Управление safety settings для администраторов"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем, является ли пользователь администратором
        if user_id not in ADMIN_IDS:
            await message.reply(
                "❌ *Доступ запрещен*\n"
                "Эта команда доступна только адмиистраторам.",
                parse_mode="Markdown"
            )
            return

        # Сздаем клавиатуру
        keyboard = InlineKeyboardBuilder()
        
        # Проверяем текущий статус safety settings
        is_disabled = user_id in DISABLED_SAFETY_USERS
        
        keyboard.add(InlineKeyboardButton(
            text=" Отключить ограничения" if not is_disabled else "🔒 Включить ограничения",
            callback_data="toggle_safety"
        ))

        status_text = (
            "🛡 *Управление ограничениями безопасности*\n\n"
            f"Текущий статус: {'🔓 Отключены' if is_disabled else '🔒 Вкючены'}\n\n"
            "⚠️ *Внимани:* Отключение ограничений может привести к генераии "
            "нежелательного кнтента. Используйте с осторожностью!"
        )

        await message.reply(
            status_text,
            reply_markup=keyboard.as_markup(),
            parse_mode="Markdown"
        )

    except Exception as e:
        logger.error(f"Error in safety_handler: {e}")
        await message.reply("Произошла ошика при управлении настройками безоасности.")

async def setup():
    """Аснхрнная нстройа бота"""
    try:
        # Инциализируем базу даных
        init_db()
        logger.info("Бза данных иницилизирована")
        
        # Устанавливам оманд бота
        commands = setup_bot_commands()
        await bot.set_my_commands(commands)
        logger.info("Команды бота установлены")
        
    except Exception as e:
        logger.error(f"Ошибка при настройке ота: {e}")
        raise

async def main() -> None:
    """Основная функция запуска бота"""
    try:
        # Инициализируем бота с правильными настройками
        bot = Bot(
            token=TELEGRAM_TOKEN,
            default=DefaultBotProperties(
                parse_mode=ParseMode.MARKDOWN_V2
            )
        )
        
        # Создаем диспетчер
        dp = Dispatcher(storage=MemoryStorage())
        
        # Настраиваем middleware
        setup_middlewares(dp)
        
        # Подключаем роутер
        dp.include_router(router)
        
        # Устанавливаем команды бота
        await bot.set_my_commands(setup_bot_commands())
        
        # Запускаем поллинг
        try:
            logger.info("Bot started")
            await dp.start_polling(
                bot,
                allowed_updates=dp.resolve_used_update_types(),
                close_bot_session=True
            )
        except Exception as e:
            logger.error(f"Polling error: {e}")
        finally:
            if bot.session:
                await bot.session.close()
            
    except Exception as e:
        logger.error(f"Startup error: {e}")
        raise

# Добавим словарь для хранения веменных изображений
temp_image_storage = {}  # user_id: {"images": [], "media_group_id": str, "waiting_prompt": bool}

@router.message(Command("analyze"))
async def analyze_handler(message: Message):
    """Обработчик команды анализа изображений"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем, есть ли изобажения для анализа
        if user_id not in user_media or not user_media[user_id]["images"]:
            await message.reply(
                "❌ У вас нет сохраненных изображений для анализа.\n"
                "Сначала отправьте одно или несколько изображений."
            )
            return
            
        images = user_media[user_id]["images"]
        
        # Отправляем сообщение об обработке
        processing_msg = await message.reply("🔄 Анализирую изображения...")
        
        try:
            # Создаем модель для обработки изображений
            model = genai.GenerativeModel('gemini-pro-vision')
            
            # Формируем запрос для анализа
            prompt = """Пожалуйста, проанализируй эти изоражения:
            1. Опиши что изображено на каждой фотографии
            2. Укажи основные детали и особенности
            3. Есл изображения связаны между собой - укажи как именно
            4. Сделай общий вывод
            """
            
            # Создаем контент для запроса
            parts = [{"text": prompt}]
            for img in images:
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                parts.append({"image": img})
            
            # Получаем ответ от модели
            response = model.generate_content(parts)
            
            if not response or not hasattr(response, 'text'):
                raise ValueError("Получен пустой ответ от модели")
            
            # Отправяем результат анализа
            await processing_msg.delete()
            await send_long_message(message, response.text)
            
            # Очищаем данные пользователя после анализа
            del user_media[user_id]
            
        except Exception as e:
            logger.error(f"Ошибка при анализе изобажений: {e}")
            await processing_msg.edit_text(
                "❌ Произошла ошиба при анализе изображений. Попробуйте еще раз."
            )
            
    except Exception as e:
        logger.error(f"Ошибка в analyze_handler: {e}")
        await message.reply(
            "❌ Произошла ошибка при выполнении команды анализа."
        )

@router.message(Command("done"), flags={"command": True})
async def done_handler(message: Message):
    """Завершает сбор изображений и запрашивает промпт"""
    try:
        user_id = str(message.from_user.id)
        if user_id not in temp_image_storage or not temp_image_storage[user_id]["images"]:
            await message.reply(
                "❌ *Нет изображений для анализа*\n\n"
                "Снаала отправьте изображения, используя команду /analyze",
                parse_mode="Markdown"
            )
            return
            
        temp_image_storage[user_id]["waiting_prompt"] = True
        
        # Формируем сводку всех изображений
        summary = "*📸 Загруженне изображения:*\n\n"
        for i, desc in enumerate(temp_image_storage[user_id]["descriptions"], 1):
            summary += f"*Изображене {i}:*\n{desc}\n\n"
            
        await message.reply(
            f"{summary}\n"
            "📝 Теперь отправьте ваш вопрос об эих изображениях\n"
            "Например:\n"
            "- Сравни эти изображния\n"
            "- Найди общие элементы\n"
            "- Какие различия между ним?\n"
            "- Что объединяет эти изображения?",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в done_handler: {e}")
        await message.reply("Произошла ошибка при обработе команды")

@router.message(F.text & ~F.text.startswith('/'))
async def message_handler(message: Message):
    """Обработчик тестовых сообщений"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем, ожидаем ли промпт для анализа изображений
        if user_id in temp_image_storage and temp_image_storage[user_id]["waiting_prompt"]:
            images = temp_image_storage[user_id]["images"]
            prompt = message.text
            
            if user_id not in chat_sessions:
                chat_sessions[user_id] = ChatSession("gemini-pro-vision", user_id)
            session = chat_sessions[user_id]
            
            try:
                response = await session.process_images(images, prompt)
                
                save_message(user_id, "user", f"[Анализ изображений] {prompt}")
                save_message(user_id, "assistant", response)
                
                api_tracker.track_request(user_id)
                
                # Используе новую функцию для отправки длинных соощений
                await send_long_message(message, response)
                
            except Exception as e:
                logger.error(f"шибка при анализе изображений: {e}")
                await message.reply(
                    "Произошла ошибка при аналзе изображений. "
                    "Попробуйте умеьшить количество изображений или изменить запрос."
                )
            finally:
                del temp_image_storage[user_id]
            
        else:
            await process_text_message(message)
            
    except Exception as e:
        logger.error(f"Ошибка в message_handler: {e}")
        await message.reply("Произошла ошибка при обработке сообщения")

# После констант и перед классами добавим функции

async def handle_single_photo(message: Message):
    """Обработчик одиночных фотографий"""
    try:
        user_id = str(message.from_user.id)
        
        can_request, error_message = await api_tracker.can_make_request(user_id)
        if not can_request:
            await message.reply(error_message)
            return

        photo = message.photo[-1]
        temp_dir = tempfile.mkdtemp()
        
        try:
            file_info = await bot.get_file(photo.file_id)
            file_path = os.path.join(temp_dir, f"image_{photo.file_id}.jpg")
            await bot.download_file(file_info.file_path, file_path)
            
            with Image.open(file_path) as img:
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                processed_image = img.copy()

            if user_id not in chat_sessions:
                chat_sessions[user_id] = ChatSession("gemini-pro-vision", user_id)  # Используем модель с поддержкой изображений
            session = chat_sessions[user_id]

            response_text = await session.process_images(
                [processed_image],
                message.caption
            )

            save_message(user_id, "user", f"[Изображение] {message.caption}")
            save_message(user_id, "assistant", response_text)
            
            api_tracker.track_request(user_id)
            
            await send_long_message(message, response_text)

        except Exception as e:
            logger.error(f"Ошибка при обработке изображения: {e}")
            await message.reply("Произошла ошибка при обработке изображения")
        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                
    except Exception as e:
        logger.error(f"Критическая ошибка в handle_single_photo: {e}")
        await message.reply("Произошла критическая ошибка при обработке фотографии")

async def process_text_message(message: Message):
    """Обработчик текстовых сообщений"""
    try:
        if not message.text:
            return

        user_id = str(message.from_user.id)
        
        can_request, error_message = await api_tracker.can_make_request(user_id)
        if not can_request:
            await message.reply(error_message)
            return
            
        if user_id not in chat_sessions:
            chat_sessions[user_id] = ChatSession(api_tracker.current_model)
            
        session = chat_sessions[user_id]
        
        # Получаем ответ от моели
        try:
            response = await session.send_message(message.text)
        except Exception as model_error:
            logger.error(f"Ошибка модели: {model_error}")
            await message.reply("Ошибка при олучении ответа от модели")
            return
            
        if not response:
            await message.reply("Получен пустой ответ от модели")
            return
            
        # Сохраняем сообщения
        save_message(user_id, "user", message.text)
        save_message(user_id, "assistant", response)
        
        # Отслеживаем использовани
        api_tracker.track_request(user_id)
        
        # Отправляем ответ
        await send_long_message(message, response, parse_mode="Markdown")
        
    except Exception as e:
        logger.error(f"Ошибка в process_text_message: {e}")
        await message.reply(
            "Произошла ошибка ри обработке сообщения. "
            "Попробуйте еще раз или измените запрос."
        )

# Обновляем костанты ля работы с сообщениями
MAX_MESSAGE_LENGTH = 3000  # Еще уменьшим для гарантии
MIN_CHUNK_SIZE = 300  # меньшим минимальный размер части
MAX_PARTS = 10  # Максимальное количество частей

async def split_text_safely(text: str) -> List[str]:
    """Безопасное разделение текста на части"""
    if not text:
        return []
        
    if len(text) <= MAX_MESSAGE_LENGTH:
        return [text]

    # Удалем лишни пробелы и перенсы строк
    text = '\n'.join(line.strip() for line in text.split('\n'))
    
    parts = []
    while text and len(parts) < MAX_PARTS:
        if len(text) <= MAX_MESSAGE_LENGTH:
            parts.append(text)
            break
            
        # Ищем место для разделения
        split_index = MAX_MESSAGE_LENGTH
        
        # Проуем найти конец предложния
        dot_index = text[:split_index].rfind('. ')
        if dot_index > MIN_CHUNK_SIZE:
            split_index = dot_index + 1
        else:
            # Пробуем найти конец абзац
            nl_index = text[:split_index].rfind('\n')
            if nl_index > MIN_CHUNK_SIZE:
                split_index = nl_index
            else:
                # Побуем найти пробел
                space_index = text[:split_index].rfind(' ')
                if space_index > MIN_CHUNK_SIZE:
                    split_index = space_index
        
        # Дбавляем часть и обрезаем текст
        parts.append(text[:split_index].strip())
        text = text[split_index:].strip()
        
    # Если текст остался, добавляем сообщение о превышении
    if text and len(parts) >= MAX_PARTS:
        parts.append("...(текст слишком длинный, показана только часть)...")
        
    return parts

def escape_markdown_v2(text: str) -> str:
    """
    Экранирование специальных символов для Markdown V2
    """
    escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    return ''.join(f'\\{c}' if c in escape_chars else c for c in text)

async def send_long_message(message: Message, text: str, **kwargs) -> None:
    """
    Отправка длинного сообщения с разбивкой на части
    """
    try:
        # Разбиваем текст на части
        if len(text) <= MAX_MESSAGE_LENGTH:
            # Разбираем текст на части для форматирования
            parts = []
            current = ""
            is_bold = False
            
            # Разбираем текст посимвольно
            i = 0
            while i < len(text):
                # Проверяем на ** или *
                if text[i:i+2] == '**' or text[i] == '*':
                    # Сохраняем накопленный текст
                    if current:
                        parts.append(Bold(current) if is_bold else current)
                        current = ""
                    is_bold = not is_bold
                    i += 2 if text[i:i+2] == '**' else 1
                    continue
                
                # Добавляем обычный текст
                current += text[i]
                i += 1
                
            # Добавляем оставшийся текст
            if current:
                parts.append(Bold(current) if is_bold else current)
                
            # Создаем форматированный текст
            content = Text(*parts)
            
            try:
                await message.answer(**content.as_kwargs())
            except TelegramBadRequest as e:
                logger.error(f"Ошибка форматирования: {e}")
                await message.answer(text=text, parse_mode=None)
            return

        # Разбиваем длинный текст на части
        parts = text.split("\n\n")
        for part in parts:
            # Форматируем каждую часть отдельно
            formatted_parts = []
            current = ""
            is_bold = False
            
            i = 0
            while i < len(part):
                if part[i:i+2] == '**' or part[i] == '*':
                    if current:
                        formatted_parts.append(Bold(current) if is_bold else current)
                        current = ""
                    is_bold = not is_bold
                    i += 2 if part[i:i+2] == '**' else 1
                    continue
                
                current += part[i]
                i += 1
                
            if current:
                formatted_parts.append(Bold(current) if is_bold else current)
                
            content = Text(*formatted_parts)
            
            try:
                await message.answer(**content.as_kwargs())
                await asyncio.sleep(0.5)
            except TelegramBadRequest as e:
                logger.error(f"Ошибка форматирования части: {e}")
                await message.answer(text=part, parse_mode=None)

    except Exception as e:
        logger.error(f"Ошибка отправки сообщения: {e}")
        await message.answer(text=text, parse_mode=None)

# Добавим константы ля форматированя
MARKDOWN_CHARS = ['*', '_', '`', '[', ']', '.', '(', ')', '#', '+', '-', '=', '|', '{', '}', '>', '<', '~']
MAX_HISTORY_LENGTH = 20

def escape_markdown(text: str) -> str:
    """Экранирование специальных символов Markdown V2"""
    for char in MARKDOWN_CHARS:
        text = text.replace(char, f"\\{char}")
    return text

def format_text_safely(text: str) -> str:
    """Беопасное форматирование текста для Markdown V2"""
    if not text:
        return ""
    
    lines = text.split('\n')
    formatted_lines = []
    in_code_block = False
    
    # Символы, которые нужно экранировать
    escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    
    for line in lines:
        line = line.strip()
        if not line:
            formatted_lines.append(line)
            continue

        # Обработка блоков кода
        if line.startswith('```'):
            in_code_block = not in_code_block
            formatted_lines.append(line)
            continue
            
        if in_code_block:
            formatted_lines.append(line)
            continue

        # Экранируем специальные символы в обычном тексте
        escaped_line = ""
        is_formatting = False
        i = 0
        while i < len(line):
            # Проверяем форматирование
            if line[i:i+2] == '**':
                escaped_line += '**'
                is_formatting = not is_formatting
                i += 2
                continue
                
            if line[i] == '*':
                escaped_line += '*'
                is_formatting = not is_formatting
                i += 1
                continue
                
            # Экранируем специальные символы только вне форматирования
            if not is_formatting and line[i] in escape_chars:
                escaped_line += f'\\{line[i]}'
            else:
                escaped_line += line[i]
            i += 1
            
        line = escaped_line

        # Форматируем списки
        if line.startswith('•'):
            line = f"• {line[2:]}"
        elif line.startswith('1️⃣'):
            line = f"1️⃣ {line[2:]}"
        elif line.startswith('│'):
            line = f"│ {line[2:]}"
            
        formatted_lines.append(line)
    
    return '\n'.join(formatted_lines)

# Добавляем улучшенные константы для форматирования сообщений
MESSAGE_FORMATTING = {
    "code_block": "```",
    "inline_code": "`",
    "bold": "*",
    "italic": "_",
    "quote": ">",
}

# Добавляем шаблоны сообщений
MESSAGE_TEMPLATES = {
    "error": "❌ *Ошибка*\n{0}",
    "success": "✅ *Успешно*\n{0}",
    "info": "ℹ️ *Информация*\n{0}",
    "warning": "⚠️ *Внимание*\n{0}",
    "file_processing": "{0}"  # Убираем лишнее форматирование, т.к. оно уже есть в результате анализа
}

def format_message(text: str, msg_type: str = "info", **kwargs) -> Text:
    """
    Форматирование текста сообщения с поддержкой различных стилей
    
    Args:
        text: Текст для форматирования
        msg_type: Тип сообщения (info, success, warning, error, code, quote)
        **kwargs: Дополнительные параметры форматирования
    """
    try:
        parts = []
        
        # Добавляем эмодзи в зависимости от типа сообщения
        prefix = {
            "info": "ℹ️ ",
            "success": "✅ ",
            "warning": "⚠️ ",
            "error": "❌ ",
        }.get(msg_type, "")
        
        if prefix:
            parts.append(prefix)
            
        # Разбираем текст на части и форматируем
        current_text = ""
        is_bold = False
        is_code = False
        i = 0
        
        while i < len(text):
            # Проверяем маркеры форматирования
            if text[i:i+2] == '**':
                if current_text:
                    parts.append(Bold(current_text) if is_bold else current_text)
                    current_text = ""
                is_bold = not is_bold
                i += 2
                continue
                
            # Проверяем маркеры кода
            elif text[i] == '`':
                if current_text:
                    parts.append(Code(current_text) if is_code else current_text)
                    current_text = ""
                is_code = not is_code
                i += 1
                continue
                
            # Обрабатываем обратные слеши
            elif text[i] == '\\':
                if i + 1 < len(text):
                    current_text += text[i+1]  # Добавляем следующий символ без экранирования
                    i += 2
                    continue
                    
            # Добавляем обычный текст
            current_text += text[i]
            i += 1
            
        # Добавляем оставшийся текст
        if current_text:
            if is_bold:
                parts.append(Bold(current_text))
            elif is_code:
                parts.append(Code(current_text))
            else:
                parts.append(current_text)
                
        return Text(*parts)
        
    except Exception as e:
        logger.error(f"Ошибка форматирования: {e}")
        return Text(text)

async def send_formatted_message(message: Message, text: str, msg_type: str = "info"):
    """Отправка отформатированного сообщения"""
    try:
        content = format_message(text, msg_type)
        await message.answer(**content.as_kwargs())
    except Exception as e:
        logger.error(f"Ошибка отправки форматированного сообщения: {e}")
        # Если возникла ошибка форматирования, отправляем без форматирования
        await message.answer(text=text, parse_mode=None)

async def process_file(message: Message, file_path: str, mime_type: str) -> str:
    """Улучшенная обработка файлов"""
    try:
        user_prompt = message.caption
        text_content = await extract_text_from_file(file_path, mime_type)
        
        if not text_content:
            return "❌ Не удалось извлечь текст из файла"

        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            chat_sessions[user_id] = ChatSession(api_tracker.current_model)
        session = chat_sessions[user_id]

        # Формируем простой промпт для анализа
        if user_prompt:
            prompt = user_prompt + "\n\n" + text_content
        else:
            prompt = (
                "рочитай и кратко опиши основные моменты текста. "
                "Используй маркированный список для ключевых пунктов.\n\n"
                "{}"
            ).format(text_content)

        # Получаем анализ от модели
        response = await session.send_message(prompt)
        
        # Формируем простой ответ
        result = (
            "📄 *{}*\n\n"
            "{}"
        ).format(
            message.document.file_name,
            response
        )
        
        return result
        
    except Exception as e:
        logger.error("Ошибка обрботки файла: {}".format(e))
        return "❌ Произошла ошибка при обработке файла"

@router.message(Command("toggle_prompt"), flags={"command": True})
async def toggle_prompt_handler(message: types.Message):
    """Включение/выключение системного промпта"""
    try:
        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            chat_sessions[user_id] = ChatSession(api_tracker.current_model)
        
        session = chat_sessions[user_id]
        is_enabled = session.toggle_system_prompt()
        
        await message.reply(
            "🔄 *Системный промпт:* {}\n\n{}".format(
                "включен" if is_enabled else "выключен",
                "Бот будет использовать системные инструкции" if is_enabled else "Бот будет отвечать напрямую на ваши сообщения"
            ),
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в toggle_prompt_handler: {e}")
        await message.reply("Произошла ошибка при изменении настроек промпта")

# Добавим новые обработчики для выбора модели

@router.callback_query(lambda c: c.data.startswith("model_page_"))
async def handle_model_pages(callback: types.CallbackQuery):
    """Обработчик переключения страниц выбора модели"""
    try:
        page = int(callback.data.split("_")[-1])
        keyboard = InlineKeyboardBuilder()
        
        if page == 1:
            models_info = MODELS_INFO["page1"]
            next_page_text = "➡️ Эксперментальные модели"
            next_page = 2
            title = "*Стандартные модели:*\n\n"
        else:
            models_info = MODELS_INFO["page2"]
            next_page_text = "⬅️ Стандартные модели"
            next_page = 1
            title = "*Экпериментальные модели:*\n\n"

        # Добавляем кнопки моделей
        for model_id, info in models_info.items():
            keyboard.add(InlineKeyboardButton(
                text=f"{info['emoji']} {info['name']}", 
                callback_data=f"model_{model_id}"
            ))

        # Добавляем кнопку переключения страницы
        keyboard.add(InlineKeyboardButton(
            text=next_page_text,
            callback_data=f"model_page_{next_page}"
        ))

        # обавляем кнопку закрытия меню
        keyboard.add(InlineKeyboardButton(
            text="❌ Закрыть",
            callback_data="close_model_menu"
        ))

        keyboard.adjust(2)  # Располагае кнопки в два столбца

        # Формируем текст сообщения
        model_text = f"🤖 *Выберите мдель:*\n\n{title}"
        
        for info in models_info.values():
            features = info.get('features', '')
            model_text += (
                f"{info['emoji']} *{info['name']}*\n"
                f"└ {info['desc']}\n"
                f"└ Скорость: {info['speed']}\n"
                f"{f'└ {features}' if features else ''}\n\n"
            )

        await callback.message.edit_text(
            model_text,
            reply_markup=keyboard.as_markup(),
            parse_mode="Markdown"
        )
        await callback.answer()

    except Exception as e:
        logger.error(f"Ошиба в handle_model_pages: {e}")
        await callback.answer("Произошла ошибка при переключении страницы", show_alert=True)

@router.callback_query(lambda c: c.data.startswith("model_"))
async def handle_model_selection(callback: types.CallbackQuery):
    """Обрабтчик выбора модели"""
    try:
        if callback.data == "close_model_menu":
            await callback.message.delete()
            await callback.answer()
            return

        model_id = callback.data.replace("model_", "")
        if model_id.startswith("page"):
            return
            
        user_id = str(callback.from_user.id)
        
        # Проверяем доступ к экспериментальным моделям
        if model_id in MODELS_INFO["page2"] and not is_admin(user_id):
            await callback.answer(
                "Эта модель доступна только администраторам", 
                show_alert=True
            )
            return

        # Проверяем лимиты для выбранной модели
        limits = MODEL_LIMITS[model_id]
        user_type = "paid" if is_admin(user_id) else "free"
        
        if limits[user_type]["rpm"] == 0:
            await callback.answer(
                "Эта модель временно недступна",
                show_alert=True
            )
            return

        # Создаем новую сессию с выбранной моделью
        chat_sessions[user_id] = ChatSession(model_id, user_id)
        
        # Получаем информацию о модели
        model_info = (MODELS_INFO["page1"].get(model_id) or 
                     MODELS_INFO["page2"].get(model_id))
        
        success_text = (
            f"✅ Выбрана модель: {model_info['emoji']} *{model_info['name']}*\n"
            f"└ {model_info['desc']}\n"
            f"└ Скорость: {model_info['speed']}"
        )
        
        if model_info.get('features'):
            success_text += f"\n└ {model_info['features']}"

        await callback.message.edit_text(
            success_text,
            parse_mode="Markdown"
        )
        await callback.answer("Модель успешно выбрана")

    except Exception as e:
        logger.error(f"шибка в handle_model_selection: {e}")
        await callback.answer(
            "Произошла ошибка при выборе модели", 
            show_alert=True
        )

@router.callback_query(lambda c: c.data == "close_model_menu")
async def close_model_menu(callback: types.CallbackQuery):
    """Обрабтчик закрытия меню выбора модели"""
    try:
        await callback.message.delete()
        await callback.answer()
    except Exception as e:
        logger.error(f"Ошибка в close_model_menu: {e}")
        await callback.answer("Ошибка при закрытии меню", show_alert=True)

# Добавим функцию проверки админа в начал файа после констант
def is_admin(user_id: str) -> bool:
    """Проверка является ли пользователь администратором"""
    return user_id in ADMIN_IDS

# Добавим новые константы для работы с историей чата
CHAT_SETTINGS = {
    "max_history_length": 1000,  # Увеличиваем максимальное количество сообщений
    "context_window": 50,        # Увеличиваем окно контекста
    "memory_threshold": 0.7,     # Порог релевантности для поиска похожих сообщений
    "max_memory_items": 20       # Увеличиваем количество "вспоминаемых" сообщений
}

class ChatMemory:
    """Класс для управления памятью чата"""
    def __init__(self):
        self.messages = []  # Все сообщения
        self.topics = {}    # Топики и связанне сообщения
        self.keywords = {}  # Клчевые слова и связанные сообщения
        
    def add_message(self, message: dict):
        """Добавление сообщения  память"""
        self.messages.append(message)
        
        # Извлекаем ключевые слова
        text = message['content'].lower()
        words = set(word for word in text.split() 
                   if len(word) > 3 and not word.isdigit())
        
        # Обовляем индекс ключевых сов
        for word in words:
            if word not in self.keywords:
                self.keywords[word] = []
            self.keywords[word].append(len(self.messages) - 1)
            
        # Ограничиваем размер стории
        if len(self.messages) > CHAT_SETTINGS["max_history_length"]:
            self.messages = self.messages[-CHAT_SETTINGS["max_history_length"]:]
            # Обновляем индексы
            self._rebuild_indices()
    
    def find_relevant_messages(self, query: str, threshold: float = None) -> List[dict]:
        """Поиск релевантных сообщений"""
        if threshold is None:
            threshold = CHAT_SETTINGS["memory_threshold"]
            
        query_words = set(query.lower().split())
        relevant_indices = set()
        
        # Ище по лючвым словам
        for word in query_words:
            if word in self.keywords:
                relevant_indices.update(self.keywords[word])
        
        # Фильтруем по релевантности
        relevant_messages = []
        for idx in relevant_indices:
            message = self.messages[idx]
            relevance = self._calculate_relevance(message['content'], query)
            if relevance >= threshold:
                relevant_messages.append({
                    **message,
                    'relevance': relevance
                })
        
        # Сортируем по релевантности и огранииваем количество
        relevant_messages.sort(key=lambda x: x['relevance'], reverse=True)
        return relevant_messages[:CHAT_SETTINGS["max_memory_items"]]
    
    def _calculate_relevance(self, text1: str, text2: str) -> float:
        """Расчет релевантнсти между двумя тестами"""
        words1 = set(text1.lower().split())
        words2 = set(text2.lower().split())
        
        intersection = words1.intersection(words2)
        union = words1.union(words2)
        
        return len(intersection) / len(union) if union else 0
    
    def _rebuild_indices(self):
        """Перестроене индексов после обезки истоии"""
        self.keywords = {}
        for i, message in enumerate(self.messages):
            text = message['content'].lower()
            words = set(word for word in text.split() 
                       if len(word) > 3 and not word.isdigit())
            
            for word in words:
                if word not in self.keywords:
                    self.keywords[word] = []
                self.keywords[word].append(i)

# Добавим словарь для хранения временых днных групп фотографий
photo_groups = {}  # user_id: {"photos": [], "timestamp": datetime, "waiting_prompt": False}

@router.message(F.text)
async def handle_text(message: Message):
    """Обработчик текстовых сообщений"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем, есть ли группа изображений, ожидающая обработи
        media_group = None
        for group_id, group in media_groups.items():
            if group["user_id"] == user_id and group["waiting_prompt"]:
                media_group = group
                media_group_id = group_id
                break

        if media_group and media_group["images"]:
            # Отправляем сообщение об обработке
            processing_msg = await message.reply("🔄 Анализирую изображения...")
            
            try:
                if user_id not in chat_sessions:
                    chat_sessions[user_id] = ChatSession("gemini-pro-vision", user_id)
                session = chat_sessions[user_id]

                # Подготавливаем изображения
                image_parts = []
                for img in media_group["images"]:
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    
                    with io.BytesIO() as bio:
                        img.save(bio, format='JPEG')
                        image_bytes = bio.getvalue()
                        image_parts.append({
                            "mime_type": "image/jpeg",
                            "data": base64.b64encode(image_bytes).decode('utf-8')
                        })

                # Формируем запрос для Gemini в правильном формате для нескольких изобржений
                request = {
                    "contents": [{
                        "parts": [
                            {"text": message.text},
                            *[{"inline_data": part} for part in image_parts]
                        ]
                    }]
                }

                # Отправляем запрос к модели
                response = await session.model.generate_content(request)

                if not response or not hasattr(response, 'text'):
                    raise ValueError("Получен пустой ответ от модели")

                response_text = response.text.strip()
                
                # Сохраняем сообщения в базу данных
                save_message(user_id, "user", f"[Группа изображений] {message.text}")
                save_message(user_id, "assistant", response_text)

                # Отправляем ответ
                await processing_msg.delete()
                await send_long_message(message, response_text)

            except Exception as e:
                logger.error(f"Ошибка при анализе изображений: {e}")
                await processing_msg.edit_text(
                    "❌ Произошла ошибка при анализе изображений. Попробуйте еще раз."
                )
            finally:
                # Очищаем данные группы
                if media_group_id in media_groups:
                    del media_groups[media_group_id]
        else:
            # Обычная обработка текстовых сообщений
            if user_id not in chat_sessions:
                chat_sessions[user_id] = ChatSession("gemini-pro", user_id)
            session = chat_sessions[user_id]
            response = await session.send_message(message.text)
            await send_long_message(message, response)

    except Exception as e:
        logger.error(f"Ошибка в handle_text: {e}")
        await message.reply("Произошла ошибка при обработке сообщения")

# В начале файла после других констант добавим:

# Константы для работы с временными файлами
TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
os.makedirs(TEMP_DIR, exist_ok=True)

# Очистка временной директории при запуске
def clear_temp_dir():
    """Очистка временной директории"""
    if os.path.exists(TEMP_DIR):
        for file in os.listdir(TEMP_DIR):
            file_path = os.path.join(TEMP_DIR, file)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                logger.error(f"Ошибка при очистке временных файлов: {e}")

# После других констант добавим:
MAX_IMAGES_PER_GROUP = 5  # Максимаьное количество изображений в грппе
MEDIA_GROUP_TIMEOUT = 60  # Таймаут для группы медиа в секундах

# Обновим структуру для хранения временных данных
media_groups = {}  # media_group_id: {"images": [], "user_id": str, "timestamp": datetime}
user_media = {}   # user_id: {"images": [], "waiting_prompt": bool, "media_group_id": str}

@router.message(F.text & ~F.text.startswith('/'))
async def handle_text_with_images(message: Message):
    """Обработчик текстоых сообщений с изображениями"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверяем, есть ли изображения для обработки
        if user_id in user_media and user_media[user_id]["images"] and user_media[user_id]["waiting_prompt"]:
            images = user_media[user_id]["images"]
            
            # Отправляем сообщение об обработке
            processing_msg = await message.reply("🔄 Анализирую изображения...")
            
            try:
                # Создаем модель для обработки изображений
                model = genai.GenerativeModel('gemini-pro-vision')
                
                # Подготавливаем изображения и текст для запроса
                parts = [{"text": message.text}]
                
                # Добавляем изображения
                for img in images:
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    parts.append({"image": img})
                
                # Получаем ответ от модели
                response = model.generate_content(parts)
                
                if not response or not hasattr(response, 'text'):
                    raise ValueError("Получен пустой ответ от модели")
                
                # Отправляем ответ
                await processing_msg.delete()
                await send_long_message(message, response.text)
                
            except Exception as e:
                logger.error(f"Ошибка при анализе изображений: {e}")
                await processing_msg.edit_text(
                    "Произошла ошибка при анализе изображений. Попробуйте еще раз."
                )
            finally:
                # Очищаем данные пользователя
                if user_id in user_media:
                    del user_media[user_id]
        
        else:
            # Обычная обработка текстовых сообщений
            await process_text_message(message)
            
    except Exception as e:
        logger.error(f"Ошибка в handle_text_with_images: {e}")
        await message.reply(
            "Произошла ошибка при обработе сообщения. Попробуйте еще раз."
        )

# Добавим периодическую очистку старых данных
async def cleanup_old_media():
    """Очистка старых медиа данных"""
    try:
        current_time = datetime.now()
        
        # Очищам старые группы медиа
        for media_group_id in list(media_groups.keys()):
            if (current_time - media_groups[media_group_id]["timestamp"]).seconds > MEDIA_GROUP_TIMEOUT:
                del media_groups[media_group_id]
        
        # Очищаем временные файлы
        if os.path.exists(TEMP_DIR):
            for user_dir in os.listdir(TEMP_DIR):
                dir_path = os.path.join(TEMP_DIR, user_dir)
                if os.path.isdir(dir_path):
                    shutil.rmtree(dir_path)
                    
    except Exception as e:
        logger.error(f"Ошибка в cleanup_old_media: {e}")

# Посе дргих констант добавим:
ALLOWED_USER_IDS = {1180375518}  # Множество разрешенных ID пользователей

# Добавим новый обработчик комнды для добавления пользователей
@router.message(Command("add_user"), flags={"command": True})
async def add_user_handler(message: types.Message):
    """Обработчик команды добавления пользователя"""
    try:
        user_id = str(message.from_user.id)
        
        # Проверем, является ли отправитель администратором
        if user_id not in ADMIN_IDS:
            await message.reply(
                "❌ *Доступ запрещен*\n"
                "Эта команда доступна только администраторам",
                parse_mode="Markdown"
            )
            return

        # Получаем ID пользователя из сообщения
        parts = message.text.split()
        if len(parts) != 2:
            await message.reply(
                "❌ *Неверный формат команды*\n"
                "Используйте: /add_user <user_id>",
                parse_mode="Markdown"
            )
            return

        try:
            new_user_id = int(parts[1])
        except ValueError:
            await message.reply(
                "❌ *Ошибка*\n"
                "ID пользователя должен быть числом",
                parse_mode="Markdown"
            )
            return

        # Добавляем пользователя в список разрешенных
        ALLOWED_USER_IDS.add(new_user_id)
        
        await message.reply(
            f"✅ *Пользователь добавлен*\n"
            f"ID: {new_user_id}",
            parse_mode="Markdown"
        )
        
        logger.info(f"Администратор {user_id} добавил пользователя {new_user_id}")

    except Exception as e:
        logger.error(f"Ошибка в add_user_handler: {e}")
        await message.reply(
            "❌ Произошла ошибка при добавлении пользователя",
            parse_mode="Markdown"
        )

# Добавим функцию проверки доступа
def is_user_allowed(user_id: int) -> bool:
    """Проверка разрешен ли доступ пользователю"""
    return user_id in ALLOWED_USER_IDS or str(user_id) in ADMIN_IDS

# Добавим функцию настройки middleware
def setup_middlewares(dp: Dispatcher) -> None:
    """Настройка middleware для бота"""
    # Добавляем только поддерживаемые middleware
    dp.message.middleware(ChatActionMiddleware())
    
    # Можно добавить свой логгер при необходимости
    @dp.message.middleware()
    async def logging_middleware(handler, event, data):
        logger.info(f"Handling message: {event.text}")
        return await handler(event, data)

# Добавим функцию для работы с историей
async def cmd_history(message: types.Message) -> None:
    """Показать историю чата"""
    try:
        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            await message.reply("У вас нет активной сессии чата")
            return
            
        session = chat_sessions[user_id]
        history = session.get_chat_history()
        
        if not history:
            await message.reply("История чата пуста")
            return
            
        # Форматируем истрию
        history_text = "*История диалога:*\n\n"
        for msg in history[-10:]:  # Показываем последние 10 сообщений
            role = "🗣" if msg["role"] == "user" else "🤖"
            content = msg["parts"][0]["text"][:100]  # Ограничиваем длину сообщения
            history_text += f"{role} {content}...\n\n"
            
        await message.reply(
            history_text,
            parse_mode=ParseMode.MARKDOWN_V2
        )
        
    except Exception as e:
        logger.error(f"Error in cmd_history: {e}")
        await message.reply("Произола ошибка при получении истории")

# В начале файла после других констант добавим:

# Константы для форматирования текста
TEXT_FORMATTING = {
    "code_block": "```{}\n{}\n```",
    "inline_code": "`{}`",
    "bold": "*{}*",
    "italic": "_{}_",
    "underline": "__{}__",
    "strikethrough": "~{}~"
}

# Шаблоны сообщений
MESSAGE_TEMPLATES = {
    "info": "ℹ️ *{}*", 
    "success": "✅ *{}*", 
    "warning": "⚠️ *{}*", 
    "error": "❌ *{}*", 
    "code": "```\n{}\n```",
    "quote": "> {}"
}

# Константы для форматирования ообщений
MESSAGE_FORMATTING = {
    "code_block": "```",
    "inline_code": "`",
    "bold": "*",
    "italic": "_",
    "quote": ">",
}

# Обновляем настройки безопасности для моделей с отключенными фильтрами
SAFETY_SETTINGS = [
    {
        "category": "HARM_CATEGORY_HARASSMENT",
        "threshold": "BLOCK_NONE"  # Полностью отключаем фильтр
    },
    {
        "category": "HARM_CATEGORY_HATE_SPEECH", 
        "threshold": "BLOCK_NONE"  # Полностью отключаем фильтр
    },
    {
        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
        "threshold": "BLOCK_NONE"  # Полностью отключаем фильтр
    },
    {
        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
        "threshold": "BLOCK_NONE"  # Полностью отключаем фильтр
    }
]

@router.message(Command("search"))
async def search_handler(message: types.Message):
    """Поиск по истории диалогов"""
    try:
        query = message.text.replace("/search", "").strip()
        if not query:
            await message.reply(
                "🔍 Укажите текст для поиска после команды:\n"
                "/search <текст для поиска>"
            )
            return

        user_id = str(message.from_user.id)
        if user_id not in chat_sessions:
            await message.reply("Нет активной сессии чата")
            return

        session = chat_sessions[user_id]
        results = await session.search_history(query)

        if not results:
            await message.reply("🔍 По вашему запросу ничего не найдено")
            return

        # Форматируем результаты
        response = ["🔍 *Результаты поиска:*\n"]
        for result in results:
            role = "👤" if result["role"] == "user" else "🤖"
            timestamp = datetime.fromisoformat(result["timestamp"]).strftime("%d.%m.%Y %H:%M")
            relevance = int(result["relevance"] * 100)
            text = result["text"][:100] + "..." if len(result["text"]) > 100 else result["text"]
            
            response.append(
                f"{role} {timestamp} (релевантность: {relevance}%)\n"
                f"{text}\n"
            )

        await send_long_message(message, "\n".join(response))

    except Exception as e:
        logger.error(f"Error in search_handler: {e}")
        await message.reply("❌ Произошла ошибка при поиске")

if __name__ == "__main__":
    try:
        # Запускаем бота
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Бот остновлен пользователем")
    except Exception as e:
        logger.error(f"Критическая ошибка: {e}")

